from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import re
import shutil
import sys
import time
import unicodedata
from collections import Counter, defaultdict, deque
from datetime import datetime
from pathlib import Path
from urllib.parse import quote, unquote, urlsplit
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from keyword_combination_engine import load_json, make_content


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
SETTINGS_PATH = ROOT / "config" / "settings.json"
PAGE_TEMPLATE = ROOT / "templates" / "school_region_page.html"
HOME_TEMPLATE = ROOT / "templates" / "school_region_home.html"
CSS_TEMPLATE = ROOT / "templates" / "school_region_site.css"
IMAGE_EXTENSIONS = {".webp", ".png", ".jpg", ".jpeg"}
METRO_NAMES = ("서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종")
SCHOOL_ALLOWED_SUFFIXES = ("과외", "수학과외", "영어과외", "국어과외", "과학과외", "내신과외")
SCHOOL_FORBIDDEN_AFTER_NAME = ("중등", "중학생", "고등", "고등학생")
PLAN_FIELDS = (
    "page_type", "scope", "지역명", "학교명", "학교급", "학년표현",
    "과목표현", "내신사용", "keyword", "slug", "시도", "행정구",
)


def stable_int(value: str) -> int:
    return int(hashlib.sha256(value.encode("utf-8")).hexdigest(), 16)


def clean(value: object) -> str:
    return unicodedata.normalize("NFC", str(value or "").strip())


def render(template: str, values: dict[str, object]) -> str:
    result = template
    for key, value in values.items():
        result = result.replace("{{" + key + "}}", str(value))
    unresolved = sorted(set(re.findall(r"\{\{([A-Za-z0-9_]+)\}\}", result)))
    if unresolved:
        raise ValueError(f"치환되지 않은 템플릿 값: {', '.join(unresolved)}")
    return result


def site_join(site_url: str, path: str) -> str:
    return site_url.rstrip("/") + "/" + path.lstrip("/")


def public_path(category: str, slug: str) -> str:
    return "/" + "/".join(quote(part, safe="") for part in (category, slug)) + "/"


def choose_preview_dir() -> Path:
    parent = ROOT / "candidate_output"
    parent.mkdir(parents=True, exist_ok=True)
    preferred = parent / "school_region_preview"
    if not preferred.exists():
        return preferred
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = parent / f"school_region_preview_{stamp}"
    suffix = 2
    while candidate.exists():
        candidate = parent / f"school_region_preview_{stamp}_{suffix}"
        suffix += 1
    return candidate


def read_excel(path: Path) -> tuple[
    list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], dict[str, int]
]:
    settings = json.loads(SETTINGS_PATH.read_text(encoding="utf-8-sig"))
    corrections = {
        str(source): str(target)
        for source, target in dict(
            settings.get("school_region_generation", {}).get(
                "school_name_corrections", {}
            )
        ).items()
    }
    workbook = load_workbook(path, read_only=True, data_only=True)
    required_sheets = {"지역별 학교", "지역 요약", "미매칭 지역"}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Excel 누락 시트: {', '.join(sorted(missing))}")

    def rows(sheet_name: str) -> list[dict[str, object]]:
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(iterator)]
        return [dict(zip(headers, values)) for values in iterator]

    raw_links = rows("지역별 학교")
    summary_rows = rows("지역 요약")
    unmatched_rows = rows("미매칭 지역")
    workbook.close()

    normalized: list[dict[str, object]] = []
    seen: set[tuple[str, str, str]] = set()
    counters = Counter()
    for source_row, raw in enumerate(raw_links, start=2):
        region = clean(raw.get("지역명"))
        source_school = clean(raw.get("학교명"))
        school = corrections.get(source_school, source_school)
        if school != source_school:
            counters["school_name_corrected"] += 1
        grade = clean(raw.get("학교급"))
        if not region or not school:
            counters["empty"] += 1
            continue
        if grade not in ("중학교", "고등학교"):
            counters["invalid_grade"] += 1
            continue
        key = (region, grade, school)
        if key in seen:
            counters["duplicate_pair"] += 1
            continue
        seen.add(key)
        normalized.append({
            "지역명": region,
            "학교급": grade,
            "학교명": school,
            "시도": clean(raw.get("시도")),
            "행정구": clean(raw.get("행정구")),
            "주소": clean(raw.get("주소")),
            "매칭기준": clean(raw.get("매칭기준")),
            "source_row": source_row,
        })
    regions = [
        {
            "지역명": clean(row.get("지역명")),
            "전체 학교 수": int(row.get("전체 학교 수") or 0),
            "중학교 수": int(row.get("중학교 수") or 0),
            "고등학교 수": int(row.get("고등학교 수") or 0),
            "매칭기준": clean(row.get("매칭기준")),
        }
        for row in summary_rows if clean(row.get("지역명"))
    ]
    unmatched = [
        {"지역명": clean(row.get("지역명")), "상태": clean(row.get("상태"))}
        for row in unmatched_rows if clean(row.get("지역명"))
    ]
    return normalized, regions, unmatched, dict(counters)


def representative_score(row: dict[str, object]) -> tuple[int, int, int, int]:
    basis = str(row["매칭기준"])
    region = str(row["지역명"])
    address = str(row["주소"])
    direct = int("직접" in basis or "지역명 일치" in basis)
    address_match = int(bool(region) and region in address)
    same_admin = int(bool(row["행정구"]) and str(row["행정구"]) in address)
    return (-direct, -address_match, -same_admin, int(row["source_row"]))


def build_school_records(links: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in links:
        groups[str(row["학교명"])].append(row)
    schools: dict[str, dict[str, object]] = {}
    for name, candidates in groups.items():
        ordered = sorted(candidates, key=representative_score)
        representative = ordered[0]
        schools[name] = {
            "학교명": name,
            "학교급": representative["학교급"],
            "대표지역": representative["지역명"],
            "시도": representative["시도"],
            "행정구": representative["행정구"],
            "주소": representative["주소"],
            "매칭기준": representative["매칭기준"],
            "관련지역": sorted({str(row["지역명"]) for row in ordered}),
            "원본연결수": len(ordered),
            "서로다른주소수": len({str(row["주소"]) for row in ordered if row["주소"]}),
        }
    return schools


def region_combinations(config: dict[str, object]) -> list[dict[str, object]]:
    subjects = [str(value) for value in config["subjects"]]
    elementary_subjects = [str(value) for value in config["elementary_subjects"]]
    if config.get("enable_elementary_korean") and "국어" not in elementary_subjects:
        elementary_subjects.append("국어")
    if config.get("enable_elementary_science") and "과학" not in elementary_subjects:
        elementary_subjects.append("과학")
    grades = ("초등", "초등학생", "중등", "중학생", "고등", "고등학생")
    rows: list[dict[str, object]] = [
        {"page_type": "region_base", "grade": "", "subject": "", "exam": False}
    ]
    rows.extend(
        {"page_type": "region_subject", "grade": "", "subject": subject, "exam": False}
        for subject in subjects
    )
    rows.extend(
        {"page_type": "region_grade", "grade": grade, "subject": "", "exam": False}
        for grade in grades
    )
    for grade in grades:
        allowed = elementary_subjects if grade.startswith("초등") else subjects
        rows.extend(
            {
                "page_type": "region_grade_subject",
                "grade": grade,
                "subject": subject,
                "exam": False,
            }
            for subject in allowed
        )
    rows.append({"page_type": "region_exam", "grade": "", "subject": "", "exam": True})
    rows.extend(
        {"page_type": "region_subject_exam", "grade": "", "subject": subject, "exam": True}
        for subject in ("수학", "영어", "국어", "과학")
    )
    rows.extend(
        {"page_type": "region_grade_exam", "grade": grade, "subject": "", "exam": True}
        for grade in ("중등", "중학생", "고등", "고등학생")
    )
    rows.extend([
        {
            "page_type": "region_grade_subject_exam",
            "grade": grade,
            "subject": subject,
            "exam": True,
        }
        for grade in ("중학생", "고등학생")
        for subject in ("수학", "영어")
    ])
    return rows


def keyword_for(entity: str, grade: str, subject: str, exam: bool) -> str:
    return f"{entity}{grade}{subject}{'내신' if exam else ''}과외"


def make_similarity_resistant_body(plan: dict[str, object]) -> str:
    """유형·slug에 따라 문단 구조와 관점을 결정적으로 달리 구성한다."""
    slug = str(plan["slug"])
    variant = int(plan.get("similarity_variant", 0))
    seed = f"{slug}:similarity-variant:{variant}"
    entity = str(plan["학교명"] or plan["지역명"])
    keyword = str(plan["keyword"])
    grade = str(plan["학년표현"])
    subject = str(plan["과목표현"])
    exam = bool(plan["내신사용"])
    school_grade = str(plan["학교급"])
    structure_names = ("상황-원인-접근-실천-점검", "문제-습관-시험-개선-유지",
                       "학교생활-진도-오답-수행-복습", "학년-난이도-시간-시험-장기",
                       "환경-성향-선택-관리-점검")
    structure = structure_names[stable_int(f"{seed}:structure") % len(structure_names)]
    perspective = ("학습 기록", "최근 평가 결과", "주간 시간표", "교과 진도표", "오답 유형")[
        stable_int(f"{seed}:perspective") % 5
    ]
    action = ("주 단위", "수업 전후", "평가 범위별", "개념 단계별", "월간 목표별")[
        stable_int(f"{seed}:action") % 5
    ]
    ending = ("다음 계획에 반영합니다", "학습 우선순위를 다시 정합니다",
              "부족한 단계를 구체적으로 보완합니다", "시험 이후의 복습까지 연결합니다",
              "스스로 점검할 수 있는 기준으로 남깁니다")[
        stable_int(f"{seed}:ending") % 5
    ]

    if plan["scope"] == "directory":
        directory_openings = (
            f"{keyword}는 {entity} 권역의 학교를 빠르게 찾기 위한 탐색 문서입니다.",
            f"{entity} 학교 탐색에서는 목록의 크기보다 학교급과 연결 지역을 정확히 구분하는 일이 먼저입니다.",
            f"{keyword} 목록은 실제 연결 자료를 기준으로 학교별 학습 페이지 진입점을 제공합니다.",
            f"{entity}의 학교 정보를 살필 때에는 학교급, 행정구, 연결 지역 순서로 범위를 좁힙니다.",
            f"{keyword}를 이용하면 지역 대표 페이지와 개별 학교 페이지를 서로 다른 경로에서 확인할 수 있습니다.",
        )
        directory_cores = (
            "중학교와 고등학교 목록을 분리하면 필요한 학년의 과목 및 내신 정보를 찾기 쉽습니다.",
            "학교 이름이 비슷하더라도 공식명과 학교급을 함께 확인해 잘못된 페이지 이동을 막습니다.",
            "광역 목록에서는 행정구 연결을 먼저 보고 세부 지역에서 실제 학교 페이지를 선택합니다.",
            "페이지를 나눈 목록은 한 화면의 링크 과밀을 줄이고 다음 탐색 단계가 분명하도록 구성합니다.",
            "대표 지역 연결은 학교 주소와 원본 매칭 기준을 따르며 임의의 지역을 추가하지 않습니다.",
        )
        directory_endings = (
            "선택한 학교에서는 수학·영어 등 과목 페이지와 내신 준비 정보를 이어서 확인합니다.",
            "목록 확인 뒤에는 지역 대표 페이지로 돌아가 주변 학교와 과목 정보를 함께 비교합니다.",
            "학교별 세부 문서에서는 공식 학교명을 유지한 채 필요한 과목의 학습 기준을 살펴봅니다.",
            "다음 목록으로 이동할 때도 같은 학교가 중복되지 않는지 학교급과 이름을 함께 점검합니다.",
            "탐색 결과는 학교 페이지와 지역 페이지 양쪽 링크를 통해 다시 확인할 수 있습니다.",
        )
        return " ".join((
            f"{keyword} 기준으로 학교급과 연결 지역을 정확하게 확인합니다.",
            directory_openings[stable_int(f"{seed}:open") % len(directory_openings)],
            directory_cores[stable_int(f"{seed}:core") % len(directory_cores)],
            f"{perspective}을 활용해 학교 선택 기준을 정리하고 {action} 탐색 결과를 검토합니다.",
            directory_endings[stable_int(f"{seed}:end") % len(directory_endings)],
            f"{structure} 순서로 확인한 내용은 {ending}.",
        ))
    if plan["scope"] == "school":
        level_focus = (
            "중학생 과목 관리와 수행평가, 고등학교 진학 전 준비"
            if school_grade == "중학교"
            else "고등 내신과 모의고사, 수시·정시를 고려한 장기 계획"
        )
        if exam:
            theme = "시험 범위와 수행평가 일정, 과목별 학습 비중, 시험 전후 복습"
        elif subject:
            subject_focus = {
                "수학": "개념 연결, 풀이 과정, 계산 실수와 오답 재풀이",
                "영어": "어휘 누적, 문장 해석, 학교 지문과 서술형 대비",
                "국어": "작품 맥락, 지문 근거, 문법과 서술형 답안 구성",
                "과학": "개념 원리, 자료 해석, 실험 과정과 단위 점검",
            }.get(subject, "교과 진도와 문제 유형")
            theme = f"{subject}의 {subject_focus}"
        else:
            theme = level_focus
        sentences = [
            f"{keyword} 계획은 {entity}의 교육과정과 평가 공지를 확인하는 데서 시작합니다.",
            f"첫 점검에서는 {perspective}을 바탕으로 {theme} 가운데 우선할 항목을 구분합니다.",
            f"{entity} 학생에게 필요한 설명은 학교명만 바꾼 일반론이 아니라 실제 진도와 평가 일정에 맞춘 순서여야 합니다.",
            f"{action} 학습에서는 이해한 내용, 혼동한 문제, 다시 풀 항목을 서로 다른 기록으로 관리합니다.",
            f"{structure} 흐름으로 상담 내용을 정리하고 결과는 {ending}.",
        ]
    else:
        if exam:
            theme = "시험 범위, 수행평가, 학교별 진도 차이, 암기와 문제풀이 비중"
        elif subject:
            subject_focus = {
                "수학": "개념 간 연결과 풀이 과정, 계산 오답",
                "영어": "어휘 누적과 문장 해석, 듣기와 서술형",
                "국어": "지문 근거와 작품 맥락, 문법 적용",
                "과학": "개념 원리와 탐구 자료, 실험 해석",
                "영수": "영어와 수학의 시간 배분과 과목별 복습",
            }.get(subject, "과목별 진도와 오답")
            theme = f"{subject} 학습의 {subject_focus}"
        elif grade:
            theme = f"{grade} 시기의 학교생활, 시험 일정, 공부 습관과 과목 우선순위"
        else:
            theme = "지역 학습환경, 학년별 흐름, 과목 선택과 학습 관리"
        sentences = [
            f"{keyword}를 살필 때는 {entity}의 {theme}을 하나의 기준으로만 판단하지 않습니다.",
            f"{perspective}에서 확인한 현재 상태를 바탕으로 학생에게 필요한 변화와 유지할 습관을 나눕니다.",
            f"{entity} 학습환경에서는 학교 일정과 이동 시간도 실제 공부 가능 시간을 결정하는 요소가 됩니다.",
            f"{action} 계획에는 핵심 문제, 실천 순서, 확인할 결과를 각각 기록해 단순 과목명 치환을 피합니다.",
            f"{structure} 구조로 학습 과정을 연결하고 점검 결과는 {ending}.",
        ]
    detail_sentences = (
        "진단표에는 정답 여부보다 풀이를 멈춘 지점과 다시 설명할 개념을 우선 기록합니다.",
        "주간 계획은 새 진도, 누적 복습, 평가 준비 시간을 서로 다른 색인으로 나누어 관리합니다.",
        "문제를 고를 때에는 맞힌 개수보다 난이도별 해결 과정과 힌트 사용 여부를 비교합니다.",
        "수업 뒤 기록은 이해·적용·설명 가능 단계로 구분해 다음 학습량을 결정하는 자료로 씁니다.",
        "시험 준비는 범위를 잘게 나눈 뒤 교과서 확인, 유형 연습, 시간 제한 점검 순으로 진행합니다.",
        "오답은 계산 실수와 개념 혼동, 조건 해석 오류를 분리해야 같은 문제가 반복되는 일을 줄일 수 있습니다.",
        "학습 시간은 긴 한 번의 수업보다 예습과 본 학습, 짧은 회상 복습으로 나눌 때 점검이 선명해집니다.",
        "설명할 수 없는 개념은 예제의 조건을 바꾸어 다시 풀고 풀이 근거를 문장으로 남깁니다.",
        "과목별 목표는 점수만 적지 않고 이번 주에 완성할 단원과 확인 문제의 범위까지 구체화합니다.",
        "평가가 끝난 뒤에는 결과를 보관하는 데 그치지 않고 다음 단원의 선행 조건과 연결해 복습합니다.",
        "학습 상담에서는 학생이 혼자 시작할 수 있는 과제와 도움이 필요한 과제를 명확히 구분합니다.",
        "진도 속도를 높이기 전 핵심 용어를 정확히 쓰는지와 기본 문제를 재현할 수 있는지부터 확인합니다.",
    )
    sentences.append(detail_sentences[stable_int(f"{seed}:detail") % len(detail_sentences)])
    # slug별 결정적 회전은 문장 무작위 섞기가 아니라 관점 배치의 차이를 만든다.
    rotation = stable_int(f"{seed}:paragraph-order") % len(sentences)
    ordered = sentences[rotation:] + sentences[:rotation]
    return " ".join(ordered)


def make_blocker_resistant_body(plan: dict[str, object]) -> str:
    """전수 검사에서 충돌한 페이지만 사용하는 넓은 결정적 문장 공간."""
    slug = str(plan["slug"])
    variant = int(plan.get("similarity_variant", 1))
    seed = f"{slug}:blocker:{variant}"
    entity_name = str(plan["학교명"] or plan["지역명"])
    keyword = str(plan["keyword"])
    subject = str(plan["과목표현"] or "종합")
    grade = str(plan["학년표현"] or "현재 학년")
    topic_map = {
        "수학": ("개념 연결", "계산 정확도", "조건 해석", "유형 분류", "오답 재풀이", "서술형 과정", "시간 배분"),
        "영어": ("어휘 회상", "문법 적용", "독해 근거", "듣기 점검", "지문 분석", "서술형 표현", "누적 복습"),
        "국어": ("지문 근거", "작품 맥락", "문법 적용", "선지 판단", "서술형 구성", "어휘 정리", "회상 복습"),
        "과학": ("개념 원리", "자료 해석", "실험 과정", "단위 확인", "탐구 서술", "오답 원인", "누적 복습"),
        "영수": ("영어 어휘", "수학 개념", "학습 시간 배분", "교차 복습", "시험 일정", "오답 기록", "주간 점검"),
        "종합": ("지역 학습환경", "학년 변화", "공부 습관", "과목 선택", "시험 준비", "시간 관리", "장기 계획"),
    }
    topics = list(topic_map.get(subject, topic_map["종합"]))
    rotate = stable_int(f"{seed}:topics") % len(topics)
    topics = topics[rotate:] + topics[:rotate]
    openings = (
        f"{keyword} 계획은 {entity_name}에서 실제로 확보할 수 있는 공부 시간부터 계산하며 시작합니다.",
        f"{entity_name}의 {keyword} 상담에서는 최근 성취도와 스스로 설명할 수 있는 범위를 먼저 구분합니다.",
        f"{keyword}를 준비할 때 첫 질문은 {entity_name} 학생이 어느 단계에서 풀이를 멈추는가입니다.",
        f"{entity_name} 학습 일정에 맞춘 {keyword}는 학교 진도와 개인 복습 속도를 따로 살펴야 합니다.",
        f"{keyword} 방향을 정하려면 {entity_name} 학생의 과제 수행 방식과 평가 기록을 함께 읽어야 합니다.",
        f"{entity_name}에서 {keyword}를 찾는 경우 현재 습관을 유지할 부분과 바꿀 부분부터 나눕니다.",
        f"{keyword} 수업 설계는 {entity_name} 학생의 정답률보다 풀이 근거가 남아 있는지를 우선 확인합니다.",
        f"{entity_name}의 시험 일정과 생활 리듬을 반영해야 {keyword} 계획이 실제 행동으로 이어집니다.",
        f"{keyword} 목표는 {entity_name}에서 사용할 교재보다 먼저 주간 학습 가능 시간을 확정해 구체화합니다.",
        f"{entity_name} 학생에게 맞는 {keyword}는 진도, 복습, 평가 준비를 서로 다른 기준으로 관리합니다.",
        f"{keyword} 상담의 출발점은 {entity_name} 학생이 혼자 시작할 수 있는 과제의 범위를 찾는 일입니다.",
        f"{entity_name}에서 진행할 {keyword}는 틀린 문제의 수가 아니라 틀린 이유의 반복 여부를 살핍니다.",
    )
    diagnoses = (
        f"진단 단계에서는 {topics[0]}과 {topics[1]}을 분리해 어느 쪽이 다음 학습을 막는지 확인합니다.",
        f"최근 기록을 {topics[1]}, {topics[2]}, {topics[3]} 기준으로 나누면 우선순위가 선명해집니다.",
        f"학생이 어려워하는 장면을 {topics[2]} 관점에서 다시 보면 단순 연습 부족과 이해 부족을 구별할 수 있습니다.",
        f"{topics[3]} 상태를 확인한 뒤 {topics[0]}으로 돌아가면 빠진 선행 단계를 찾기 쉽습니다.",
        f"첫 주에는 {topics[4]} 기록과 {topics[1]} 결과를 비교해 반복되는 오류만 추립니다.",
        f"{topics[5]}을 말로 설명하게 하고 {topics[2]} 근거를 표시하면 막연한 이해를 피할 수 있습니다.",
        f"평가 결과는 {topics[0]}, {topics[3]}, {topics[6]} 세 축으로 기록해 점수 하나로 판단하지 않습니다.",
        f"{topics[6]}이 불안정하면 새 진도보다 {topics[4]}의 회수와 간격을 먼저 조정합니다.",
        f"수업 전 확인에서는 {topics[1]}보다 {topics[5]}의 재현 가능성을 먼저 점검합니다.",
        f"{topics[2]}에서 생긴 실수가 {topics[0]} 부족인지 {topics[6]} 문제인지 구분해 기록합니다.",
        f"주간 진단표에는 {topics[3]} 달성도와 {topics[4]} 완료 시점을 별도 칸에 남깁니다.",
        f"{topics[5]} 결과를 {topics[1]} 기록과 대조하면 다음 문제의 난이도를 안정적으로 정할 수 있습니다.",
    )
    actions = (
        f"실천 계획은 {topics[0]} 확인, {topics[3]} 연습, {topics[4]} 재점검의 세 구간으로 운영합니다.",
        f"하루 학습을 {topics[1]} 훈련과 {topics[5]} 설명 시간으로 나누고 마지막에 {topics[6]}을 배치합니다.",
        f"새 문제를 풀기 전에 {topics[2]} 기준을 적고 완료 후 {topics[4]} 원인을 한 문장으로 정리합니다.",
        f"{topics[3]}별 대표 문제를 고른 다음 {topics[0]}과 연결되는 이유를 스스로 말하게 합니다.",
        f"주중에는 {topics[1]}을 짧게 반복하고 주말에는 {topics[5]}을 시간 제한 안에서 확인합니다.",
        f"{topics[6]} 일정은 한 번에 몰지 않고 {topics[0]} 학습 직후와 이틀 뒤로 나누어 배치합니다.",
        f"과제량은 {topics[2]} 성공률에 따라 조정하며 {topics[4]}가 남으면 새 진도를 늦춥니다.",
        f"{topics[5]} 답안은 근거, 과정, 결론을 구분하고 {topics[1]} 실수는 별도 목록으로 관리합니다.",
        f"수업 시작 전 {topics[0]}을 회상하고 종료 전에는 {topics[3]}의 선택 기준을 다시 적습니다.",
        f"{topics[4]} 문제는 정답을 외우지 않고 조건을 바꾸어 {topics[2]} 과정을 다시 수행합니다.",
        f"{topics[6]} 시간을 확보하기 위해 {topics[1]} 과제는 핵심 문항과 확인 문항으로 나눕니다.",
        f"{topics[3]} 진도표 옆에 {topics[5]} 점검일을 표시해 시험 직전의 과부하를 줄입니다.",
    )
    examples = (
        f"예를 들어 {topics[0]}은 맞았지만 {topics[2]}에서 멈췄다면 문제 수보다 조건 표시 연습을 늘립니다.",
        f"{topics[1]} 실수가 반복되는 날에는 난이도를 올리지 않고 같은 원리를 다른 형식으로 확인합니다.",
        f"시험 범위가 넓을 때는 {topics[3]} 단위를 작게 나누고 완료한 범위만 다음 복습표로 넘깁니다.",
        f"과제 시작이 늦어지는 학생은 {topics[6]} 분량을 십 분 안에 끝낼 수 있는 크기로 줄여 시작 장벽을 낮춥니다.",
        f"{topics[5]} 표현이 막히면 모범 답안을 베끼기보다 핵심 근거 세 개를 먼저 배열합니다.",
        f"{topics[4]}가 누적될 때는 틀린 날짜보다 다시 해결한 날짜를 기록해 회복 속도를 확인합니다.",
        f"{topics[2]} 판단이 흔들리면 문제의 수치나 소재를 바꾸어 같은 판단 기준이 유지되는지 봅니다.",
        f"{topics[0]} 설명은 가능하지만 적용이 느리면 제한 시간을 두기 전에 풀이 순서를 고정합니다.",
        f"{topics[3]}별 점수가 들쭉날쭉하면 가장 약한 유형 하나만 골라 원인과 행동을 연결합니다.",
        f"{topics[1]} 정확도가 낮은 주에는 전체 과제량을 줄이고 검산 또는 재확인 절차를 추가합니다.",
        f"{topics[6]}이 밀리면 새 내용을 더하지 않고 이전 학습을 짧게 회상하는 날을 먼저 확보합니다.",
        f"{topics[5]} 결과가 모호하면 학생이 사용한 근거와 교과서 표현을 나란히 놓고 차이를 찾습니다.",
    )
    checks = (
        f"점검표에는 {topics[0]} 이해, {topics[4]} 해결, {topics[6]} 완료 여부를 서로 다른 항목으로 남깁니다.",
        f"다음 수업의 난이도는 {topics[2]} 정확도와 {topics[5]} 설명 가능 여부를 함께 보고 정합니다.",
        f"주간 평가는 {topics[1]} 결과보다 {topics[3]} 선택 이유가 일관적인지를 우선 확인합니다.",
        f"{topics[4]}를 다시 틀리면 풀이를 외웠는지 살피고 {topics[0]} 질문으로 되돌아갑니다.",
        f"월말에는 {topics[6]} 간격과 {topics[2]} 성공률을 비교해 학습 주기를 다시 계산합니다.",
        f"학생이 {topics[5]}을 혼자 완성하면 {topics[3]} 범위를 넓히고 그렇지 않으면 예시 수를 줄입니다.",
        f"점검 결과는 완료·보완·재설명 세 단계로 표시하고 {topics[1]} 문제는 다음 주 첫 과제로 옮깁니다.",
        f"{topics[0]} 회상이 느려진 경우 {topics[6]} 시간을 늘리되 새 문제의 양은 그대로 두지 않습니다.",
        f"{topics[3]} 성취는 문항 수가 아니라 대표 유형을 다른 조건에서도 해결했는지로 판단합니다.",
        f"{topics[2]} 기준을 말로 설명하지 못하면 정답 여부와 관계없이 보완 목록에 포함합니다.",
        f"시험 뒤에는 {topics[4]} 원인을 분류하고 {topics[5]} 답안을 다음 단원 학습과 연결합니다.",
        f"{topics[1]} 개선이 확인되면 검산 횟수보다 스스로 오류를 발견한 시점을 기록합니다.",
    )
    endings = (
        "이 기록을 다음 주 계획에 반영하면 진도보다 학습 과정의 변화를 지속해서 확인할 수 있습니다.",
        "마지막에는 학생이 스스로 유지할 행동 한 가지를 정해 수업 밖에서도 계획이 이어지게 합니다.",
        "평가가 끝난 뒤에도 같은 점검표를 사용하면 단기 준비와 장기 습관을 함께 관리할 수 있습니다.",
        "완료한 범위와 다시 볼 범위를 분리해 두면 다음 상담에서 목표를 처음부터 다시 세우지 않아도 됩니다.",
        "결과는 점수만 비교하지 않고 설명 능력과 복습 실행률을 함께 보며 다음 단계를 결정합니다.",
        "학습량을 늘리는 결정은 기록이 안정된 뒤에 하며, 먼저 반복 오류가 줄었는지 확인합니다.",
        "학생이 계획의 이유를 이해하도록 설명하면 외부 관리가 줄어도 같은 학습 흐름을 유지할 수 있습니다.",
        "다음 시험에서는 기존 기록을 출발점으로 사용해 준비 기간과 과목별 비중을 더 정확히 배분합니다.",
        "보완할 항목은 한꺼번에 늘리지 않고 가장 영향이 큰 한 가지부터 다음 주 목표로 옮깁니다.",
        "수업과 자습의 역할을 구분해 두면 시간이 부족한 주에도 핵심 복습이 빠지는 일을 줄일 수 있습니다.",
        "최종 점검에서는 학생이 직접 변화한 점을 설명하게 해 계획의 실행 가능성을 확인합니다.",
        "이 과정을 누적하면 단원과 시험이 달라져도 스스로 학습 순서를 조정하는 기준이 남습니다.",
    )
    banks = (openings, diagnoses, actions, examples, checks, endings)
    selected = [
        bank[stable_int(f"{seed}:bank:{index}") % len(bank)]
        for index, bank in enumerate(banks)
    ]
    structures = (
        (0, 1, 2, 3, 4, 5), (0, 2, 1, 4, 3, 5), (1, 0, 3, 2, 4, 5),
        (3, 1, 0, 2, 4, 5), (0, 3, 2, 1, 5, 4), (1, 2, 0, 4, 3, 5),
        (2, 0, 1, 3, 5, 4), (0, 1, 3, 4, 2, 5), (3, 0, 2, 4, 1, 5),
        (1, 3, 0, 2, 5, 4), (2, 1, 3, 0, 4, 5), (0, 4, 1, 3, 2, 5),
    )
    order = structures[stable_int(f"{seed}:structure") % len(structures)]
    return " ".join(selected[index] for index in order)


def generation_phase(plan: dict[str, object]) -> str:
    if plan["scope"] == "school":
        return "school"
    if plan["scope"] == "directory":
        return "directory"
    if plan["page_type"] in ("region_subject", "region_subject_exam"):
        return "subject"
    if plan["page_type"] in (
        "region_grade", "region_grade_subject", "region_grade_exam",
        "region_grade_subject_exam",
    ):
        return "grade"
    return "region"


def plan_pages(
    regions: list[dict[str, object]],
    links: list[dict[str, object]],
    schools: dict[str, dict[str, object]],
    config: dict[str, object],
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    link_by_region: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in links:
        link_by_region[str(row["지역명"])].append(row)
    plans: list[dict[str, object]] = []
    region_specs = region_combinations(config)
    for region_row in sorted(regions, key=lambda row: str(row["지역명"])):
        region = str(region_row["지역명"])
        linked = link_by_region.get(region, [])
        province = str(linked[0]["시도"]) if linked else ""
        district = str(linked[0]["행정구"]) if linked else ""
        for combo in region_specs:
            keyword = keyword_for(
                region, str(combo["grade"]), str(combo["subject"]), bool(combo["exam"])
            )
            plans.append({
                "page_type": combo["page_type"],
                "scope": "region",
                "지역명": region,
                "학교명": "",
                "학교급": "",
                "학년표현": combo["grade"],
                "과목표현": combo["subject"],
                "내신사용": bool(combo["exam"]),
                "keyword": keyword,
                "slug": keyword.replace(" ", ""),
                "시도": province,
                "행정구": district,
            })
    validation_rows: list[dict[str, object]] = []
    for school_name, school in sorted(schools.items()):
        for suffix in SCHOOL_ALLOWED_SUFFIXES:
            subject = suffix.removesuffix("과외") if suffix not in ("과외", "내신과외") else ""
            exam = suffix == "내신과외"
            keyword = school_name + suffix
            plan = {
                "page_type": (
                    "school_base" if suffix == "과외"
                    else "school_exam" if exam else "school_subject"
                ),
                "scope": "school",
                "지역명": school["대표지역"],
                "학교명": school_name,
                "학교급": school["학교급"],
                "학년표현": "",
                "과목표현": subject,
                "내신사용": exam,
                "keyword": keyword,
                "slug": keyword.replace(" ", ""),
                "시도": school["시도"],
                "행정구": school["행정구"],
            }
            plans.append(plan)
            invalid = any(
                school_name + token in keyword for token in SCHOOL_FORBIDDEN_AFTER_NAME
            )
            validation_rows.append({
                "학교명": school_name,
                "학교급": school["학교급"],
                "keyword": keyword,
                "slug": plan["slug"],
                "허용조합": suffix in SCHOOL_ALLOWED_SUFFIXES,
                "학교급중복표현": invalid,
                "validation_note": "정상" if not invalid else "학교명 뒤 학교급 표현",
            })

    page_size = int(config["school_directory_page_size"])
    for metro in METRO_NAMES:
        if not any(str(row["지역명"]) == metro for row in regions):
            continue
        for grade in ("중학교", "고등학교"):
            matching = [
                school for school in schools.values()
                if school["학교급"] == grade
                and (
                    school["대표지역"] == metro
                    or str(school["시도"]).startswith(metro)
                    or metro in str(school["시도"])
                )
            ]
            for page_number in range(1, max(1, math.ceil(len(matching) / page_size)) + 1):
                suffix = "" if page_number == 1 else str(page_number)
                keyword = f"{metro}{grade}전체{suffix}"
                plans.append({
                    "page_type": "metro_school_directory",
                    "scope": "directory",
                    "지역명": metro,
                    "학교명": "",
                    "학교급": grade,
                    "학년표현": "",
                    "과목표현": "",
                    "내신사용": False,
                    "keyword": keyword,
                    "slug": keyword,
                    "시도": metro,
                    "행정구": "",
                    "directory_page": page_number,
                })

    duplicate_rows: list[dict[str, object]] = []
    for plan in plans:
        plan["similarity_enhanced"] = bool(
            config.get("enable_similarity_resistant_body", False)
        )
    variant_file = config.get("priority1_blocker_variant_file")
    if variant_file:
        variant_path = ROOT / str(variant_file)
        if variant_path.is_file():
            variant_map = json.loads(variant_path.read_text(encoding="utf-8-sig"))
            for plan in plans:
                value = variant_map.get(str(plan["slug"]))
                if value is not None:
                    plan["similarity_variant"] = int(value)
                    plan["blocker_variant"] = True
    repair_variant_file = config.get("production_full_repair_variant_file")
    if repair_variant_file:
        repair_variant_path = ROOT / str(repair_variant_file)
        if repair_variant_path.is_file():
            repair_variants = json.loads(
                repair_variant_path.read_text(encoding="utf-8-sig")
            )
            for plan in plans:
                value = repair_variants.get(str(plan["slug"]))
                if value is not None:
                    plan["similarity_variant"] = int(value)
                    plan["blocker_variant"] = True
    counts = Counter(str(plan["slug"]) for plan in plans)
    duplicate_slugs = {slug for slug, count in counts.items() if count > 1}
    if duplicate_slugs:
        kept: set[str] = set()
        unique_plans: list[dict[str, object]] = []
        for plan in plans:
            slug = str(plan["slug"])
            if slug in kept:
                duplicate_rows.append({
                    "slug": slug,
                    "page_type": plan["page_type"],
                    "지역명": plan["지역명"],
                    "학교명": plan["학교명"],
                    "처리": "후속 중복 계획 제외",
                })
                continue
            kept.add(slug)
            unique_plans.append(plan)
        plans = unique_plans
    return plans, duplicate_rows, validation_rows


def select_sample(
    plans: list[dict[str, object]],
    regions: list[dict[str, object]],
    schools: dict[str, dict[str, object]],
    config: dict[str, object],
) -> tuple[list[dict[str, object]], list[str], list[str]]:
    region_names = sorted(str(row["지역명"]) for row in regions)
    metros = [name for name in METRO_NAMES if name in region_names][:2]
    dong = [name for name in region_names if name.endswith("동") and name not in metros]
    selected_regions = metros + dong[:5]
    selected_regions.extend(
        name for name in region_names
        if name not in selected_regions and len(selected_regions) < int(config["sample_region_count"])
    )

    region_school_names: set[str] = set()
    for school in schools.values():
        if any(region in selected_regions for region in school["관련지역"]):
            region_school_names.add(str(school["학교명"]))
    middle = sorted(
        (
            name for name, school in schools.items()
            if school["학교급"] == "중학교"
        ),
        key=lambda name: (name not in region_school_names, stable_int(f"sample-middle:{name}")),
    )[: int(config["sample_middle_school_count"])]
    high = sorted(
        (
            name for name, school in schools.items()
            if school["학교급"] == "고등학교"
        ),
        key=lambda name: (name not in region_school_names, stable_int(f"sample-high:{name}")),
    )[: int(config["sample_high_school_count"])]
    selected_schools = middle + high

    plan_by_key = {
        (
            plan["scope"], plan["지역명"], plan["학교명"], plan["학년표현"],
            plan["과목표현"], bool(plan["내신사용"]), plan["page_type"],
        ): plan for plan in plans
    }
    sample: list[dict[str, object]] = []
    per_region_specs = [
        ("region_base", "", "", False),
        *[("region_subject", "", subject, False) for subject in ("수학", "영어", "국어", "과학", "영수")],
        ("region_grade", "초등학생", "", False),
        ("region_grade", "중학생", "", False),
        ("region_grade", "고등학생", "", False),
        ("region_grade_subject", "초등학생", "수학", False),
        ("region_grade_subject", "중학생", "영어", False),
        ("region_grade_subject", "고등학생", "과학", False),
        ("region_exam", "", "", True),
        ("region_subject_exam", "", "수학", True),
        ("region_grade_exam", "중학생", "", True),
        ("region_grade_subject_exam", "고등학생", "영어", True),
    ]
    for region in selected_regions:
        for page_type, grade, subject, exam in per_region_specs:
            key = ("region", region, "", grade, subject, exam, page_type)
            if key in plan_by_key:
                sample.append(plan_by_key[key])
    for school_name in selected_schools:
        sample.extend(
            plan for plan in plans
            if plan["scope"] == "school" and plan["학교명"] == school_name
        )
    for metro in metros:
        sample.extend(
            plan for plan in plans
            if plan["scope"] == "directory"
            and plan["지역명"] == metro
            and int(plan.get("directory_page", 1)) == 1
        )
    sample = list({str(plan["slug"]): plan for plan in sample}.values())
    limit = int(config["sample_page_limit"])
    if len(sample) > limit:
        raise ValueError(f"샘플 계획이 제한 {limit}개를 초과했습니다: {len(sample)}")
    return sample, selected_regions, selected_schools


def write_csv(path: Path, fields: list[str] | tuple[str, ...], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_plan_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "샘플 페이지 계획"
    sheet.append(PLAN_FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in PLAN_FIELDS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="355070")
    for index, field in enumerate(PLAN_FIELDS, start=1):
        width = max(len(str(field)), *(
            len(str(row.get(field, ""))) for row in rows
        ))
        sheet.column_dimensions[get_column_letter(index)].width = min(width + 3, 45)
    workbook.save(path)


def create_title(
    plan: dict[str, object],
    pools: dict[str, list[dict[str, str]]],
    used_titles: set[str],
    title_min: int,
    title_max: int,
) -> tuple[dict[str, object], int]:
    if plan["scope"] == "directory":
        keyword = str(plan["keyword"])
        title = f"{keyword} 연결 학교와 지역별 과외 학습 및 내신 준비 탐색 기준을 자세히 정리합니다"
        if len(title) < title_min:
            title += " 학교별 과목 선택 방향도 함께 살펴봅니다"
        if len(title) > title_max:
            title = f"{keyword} 연결 학교와 지역별 과외 학습 탐색 기준을 정리합니다"
        if title in used_titles:
            raise ValueError(f"학교 탐색 title 중복: {plan['slug']}")
        used_titles.add(title)
        directory_content = {
            "title": title,
            "description": f"{keyword}에서 연결된 실제 학교와 지역별 과외 학습 페이지를 찾아볼 수 있습니다.",
            "body": f"{keyword}는 학교 목록 탐색용 페이지입니다. 연결된 학교별 과목 과외와 내신 학습 페이지를 확인할 수 있습니다.",
            "body_ids": [],
            "used_pools": [],
        }
        if bool(plan.get("similarity_enhanced")):
            directory_content["body"] = make_similarity_resistant_body(plan)
        return directory_content, 0
    combo = {
        "type": str(plan["page_type"]),
        "grade": str(plan["학년표현"]),
        "subject": str(plan["과목표현"]),
        "exam": bool(plan["내신사용"]),
    }
    entity = str(plan["학교명"] or plan["지역명"])
    identity = str(plan["slug"])
    retry = 0
    generated = make_content(entity, combo, identity, pools)

    def normalize_keyword(values: dict[str, object]) -> dict[str, object]:
        source_keyword = str(values["keyword"])
        exact_keyword = str(plan["keyword"])
        for field in ("title", "description", "body"):
            values[field] = str(values[field]).replace(source_keyword, exact_keyword, 1)
        title_value = str(values["title"])
        additions = (
            " 학습 방향을 함께 살펴봅니다",
            " 시험 준비와 복습 기준을 정리합니다",
            " 개념 이해와 오답 관리 방법을 살펴봅니다",
        )
        for addition in additions:
            if len(title_value) >= title_min:
                break
            title_value += addition
        values["title"] = title_value
        return values

    generated = normalize_keyword(generated)
    if bool(plan.get("blocker_variant")):
        generated["body"] = make_blocker_resistant_body(plan)
    elif bool(plan.get("similarity_enhanced")):
        generated["body"] = make_similarity_resistant_body(plan)
    while str(generated["title"]) in used_titles and retry < 100:
        retry += 1
        generated = make_content(entity, combo, f"{identity}:title:retry_{retry}", pools)
        generated = normalize_keyword(generated)
        if bool(plan.get("blocker_variant")):
            generated["body"] = make_blocker_resistant_body(plan)
        elif bool(plan.get("similarity_enhanced")):
            generated["body"] = make_similarity_resistant_body(plan)
    title = str(generated["title"])
    if not title.startswith(str(plan["keyword"])):
        raise ValueError(f"title 필수 키워드 앞부분 누락: {plan['slug']}")
    if not title_min <= len(title) <= title_max:
        raise ValueError(f"title 길이 오류({len(title)}): {plan['slug']}")
    if title in used_titles:
        raise ValueError(f"title 중복 미해결: {plan['slug']}")
    used_titles.add(title)
    return generated, retry


def generate_preview(
    output: Path,
    sample: list[dict[str, object]],
    selected_regions: list[str],
    schools: dict[str, dict[str, object]],
    links: list[dict[str, object]],
    settings: dict[str, object],
    config: dict[str, object],
) -> dict[str, object]:
    started = time.perf_counter()
    category = str(settings["page_category"])
    site_url = str(settings["site_url"]).rstrip("/")
    output.mkdir(parents=True, exist_ok=False)
    page_template = PAGE_TEMPLATE.read_text(encoding="utf-8-sig")
    home_template = HOME_TEMPLATE.read_text(encoding="utf-8-sig")
    assets_output = output / "assets"
    assets_output.mkdir(parents=True, exist_ok=True)
    shutil.copy2(CSS_TEMPLATE, assets_output / "site.css")
    pool_names = (
        "description_general", "openings", "endings", "body_general",
        "body_elementary", "body_middle", "body_high", "body_korean",
        "body_english", "body_math", "body_science", "body_english_math",
        "body_internal_exam", "title_patterns", "title_modifiers", "title_endings",
    )
    pools = {name: load_json(name) for name in pool_names}
    title_min = int(settings["content_generation"]["title_min_length"])
    title_max = int(settings["content_generation"]["title_max_length"])

    common_dir = ROOT / str(settings["image_dir"])
    thumbnail_dir = ROOT / str(settings["thumbnail_dir"])
    common_files = sorted(
        path for path in common_dir.iterdir()
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
    ) if common_dir.is_dir() else []
    thumbnails = sorted(
        path for path in thumbnail_dir.iterdir()
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
    ) if thumbnail_dir.is_dir() else []
    common = common_files[0] if common_files else None
    copied: set[Path] = set()

    def copy_asset(source: Path, configured_dir: str) -> str:
        destination = output / configured_dir / source.name
        if destination not in copied:
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, destination)
            copied.add(destination)
        return "/" + quote(configured_dir.replace("\\", "/").strip("/"), safe="/") + "/" + quote(source.name)

    common_url = copy_asset(common, str(settings["image_dir"])) if common else ""

    def image_dimensions(path: Path) -> tuple[int, int]:
        with Image.open(path) as image:
            return int(image.width), int(image.height)

    sample_by_slug = {str(plan["slug"]): plan for plan in sample}
    global_fallback_slugs = [str(plan["slug"]) for plan in sample[:60]]
    dimension_cache = {
        path: image_dimensions(path)
        for path in ([common] if common else []) + thumbnails
    }
    region_slugs: dict[str, list[str]] = defaultdict(list)
    school_slugs: dict[str, list[str]] = defaultdict(list)
    for plan in sample:
        if plan["scope"] == "region":
            region_slugs[str(plan["지역명"])].append(str(plan["slug"]))
        elif plan["scope"] == "school":
            school_slugs[str(plan["학교명"])].append(str(plan["slug"]))

    school_names_by_region: dict[str, list[str]] = defaultdict(list)
    for row in links:
        school_names_by_region[str(row["지역명"])].append(str(row["학교명"]))
    generated_records: list[dict[str, object]] = []
    used_titles: set[str] = set()
    used_descriptions: set[str] = set()
    used_bodies: set[str] = set()
    retry_count = 0

    for plan in sample:
        generated, retries = create_title(plan, pools, used_titles, title_min, title_max)
        retry_count += retries
        keyword = str(plan["keyword"])
        title = str(generated["title"])
        description = str(generated["description"])
        body = str(generated["body"])
        link_candidates: list[str] = []
        if plan["scope"] == "region":
            siblings = region_slugs[str(plan["지역명"])]
            index = siblings.index(str(plan["slug"]))
            link_candidates.extend(
                siblings[(index + offset) % len(siblings)] for offset in range(1, len(siblings))
            )
            for school_name in school_names_by_region.get(str(plan["지역명"]), []):
                if school_name in school_slugs:
                    link_candidates.append(school_slugs[school_name][0])
        elif plan["scope"] == "school":
            siblings = school_slugs[str(plan["학교명"])]
            link_candidates.extend(slug for slug in siblings if slug != plan["slug"])
            for region in schools[str(plan["학교명"])]["관련지역"]:
                bases = [
                    slug for slug in region_slugs.get(str(region), [])
                    if sample_by_slug[slug]["page_type"] == "region_base"
                ]
                link_candidates.extend(bases)
        else:
            link_candidates.extend(
                slugs[0] for name, slugs in school_slugs.items()
                if schools[name]["학교급"] == plan["학교급"]
            )
        link_candidates.extend(global_fallback_slugs)
        internal_links: list[str] = []
        for slug in link_candidates:
            if slug != plan["slug"] and slug not in internal_links:
                internal_links.append(slug)
            if len(internal_links) >= int(config["internal_link_max"]) - 5:
                break
        links_html = "<br>\n<strong>관련 학습 페이지</strong><br>\n" + " · ".join(
            f'<a href="{public_path(category, slug)}">{html.escape(str(sample_by_slug[slug]["keyword"]))}</a>'
            for slug in internal_links
        )

        image_blocks: list[str] = []
        preload = ""
        if common_url:
            common_width, common_height = dimension_cache[common]
            image_blocks.append(
                f'    <figure><img src="{common_url}" width="{common_width}" '
                f'height="{common_height}" loading="eager" decoding="async" '
                f'alt="{html.escape(keyword)}"></figure>'
            )
            preload = f'  <link rel="preload" as="image" href="{common_url}">'
        thumbnail_url = ""
        if thumbnails:
            selected = thumbnails[stable_int(str(plan["slug"])) % len(thumbnails)]
            thumbnail_url = copy_asset(selected, str(settings["thumbnail_dir"]))
            thumb_width, thumb_height = dimension_cache[selected]
            image_blocks.append(
                f'    <figure><img src="{thumbnail_url}" width="{thumb_width}" '
                f'height="{thumb_height}" loading="lazy" decoding="async" '
                f'alt="{html.escape(keyword)}"></figure>'
            )
        canonical = site_join(site_url, public_path(category, str(plan["slug"])))
        og_url = thumbnail_url or common_url
        og_meta = (
            f'  <meta property="og:image" content="{html.escape(site_join(site_url, og_url), quote=True)}">'
            if og_url else ""
        )
        twitter_meta = (
            f'  <meta name="twitter:image" content="{html.escape(site_join(site_url, og_url), quote=True)}">'
            if og_url else ""
        )
        page_html = render(page_template, {
            "language": html.escape(str(settings["language"]), quote=True),
            "title": html.escape(title),
            "description": html.escape(description, quote=True),
            "keywords": html.escape(keyword, quote=True),
            "canonical": html.escape(canonical, quote=True),
            "og_image_meta": og_meta,
            "twitter_image_meta": twitter_meta,
            "preload_image": preload,
            "background_color": html.escape(str(settings["background_color"])),
            "max_width": int(settings["max_width"]),
            "image_blocks": "\n".join(image_blocks),
            "keyword": html.escape(keyword),
            "body_html": html.escape(keyword) + "<br>\n" + html.escape(body),
            "internal_links": "\n".join(
                f'        <li><a href="{public_path(category, slug)}">'
                f'{html.escape(str(sample_by_slug[slug]["keyword"]))}</a></li>'
                for slug in internal_links
            ),
        })
        page_file = output / category / str(plan["slug"]) / "index.html"
        page_file.parent.mkdir(parents=True, exist_ok=False)
        page_file.write_text(page_html, encoding="utf-8")
        generated_records.append({
            **plan,
            "title": title,
            "description": description,
            "body": body,
            "canonical": canonical,
            "path": public_path(category, str(plan["slug"])),
            "file": str(page_file),
            "internal_links": internal_links,
        })
        used_descriptions.add(description)
        used_bodies.add(body)

    region_base_records = [
        record for record in generated_records if record["page_type"] == "region_base"
    ]
    navigation_pages: list[dict[str, str]] = []
    planned_hubs = {
        "지역찾기": "지역별 과외 찾기",
        "과목찾기": "과목별 과외 찾기",
        "수학과외찾기": "수학과외 지역 찾기",
        "영어과외찾기": "영어과외 지역 찾기",
        "국어과외찾기": "국어과외 지역 찾기",
        "과학과외찾기": "과학과외 지역 찾기",
        "영수과외찾기": "영수과외 지역 찾기",
        "학년찾기": "학년별 과외 찾기",
        "학교찾기": "학교별 과외 찾기",
        "중학교찾기": "중학교 과외 찾기",
        "고등학교찾기": "고등학교 과외 찾기",
        "전체샘플찾기": "전체 검수 페이지 찾기",
    }
    common_hub_links = [
        (public_path(category, slug), label) for slug, label in planned_hubs.items()
    ]

    def write_hub(slug: str, heading: str, description: str, links: list[tuple[str, str]]) -> None:
        unique: list[tuple[str, str]] = []
        for item in links + common_hub_links + [
            (str(record["path"]), str(record["keyword"])) for record in region_base_records
        ]:
            if item[0] != public_path(category, slug) and item not in unique:
                unique.append(item)
            if len(unique) >= int(config["internal_link_max"]) - 1:
                break
        hub_path = public_path(category, slug)
        hub_html = render(home_template, {
            "language": html.escape(str(settings["language"]), quote=True),
            "title": html.escape(f"{heading} | 공부업"),
            "description": html.escape(description, quote=True),
            "canonical": html.escape(site_join(site_url, hub_path), quote=True),
            "heading": html.escape(heading),
            "body_html": html.escape(description),
            "internal_links": "\n".join(
                f'      <a href="{href}">{html.escape(label)}</a>'
                for href, label in unique
            ),
        })
        hub_file = output / category / slug / "index.html"
        hub_file.parent.mkdir(parents=True, exist_ok=False)
        hub_file.write_text(hub_html, encoding="utf-8")
        navigation_pages.append({
            "slug": slug, "path": hub_path,
            "canonical": site_join(site_url, hub_path), "family": "navigation",
        })

    write_hub(
        "지역찾기", "지역별 과외 찾기",
        "지역 대표 페이지에서 과목·학년·내신 및 연결 학교 페이지를 탐색할 수 있습니다.",
        [(str(record["path"]), str(record["keyword"])) for record in region_base_records],
    )
    subject_hub_links = [
        (public_path(category, f"{subject}과외찾기"), f"{subject}과외 지역 찾기")
        for subject in ("수학", "영어", "국어", "과학", "영수")
    ]
    write_hub("과목찾기", "과목별 과외 찾기", "과목을 선택해 주요 지역 페이지로 이동합니다.", subject_hub_links)
    for subject in ("수학", "영어", "국어", "과학", "영수"):
        subject_records = [
            record for record in generated_records
            if record["page_type"] == "region_subject"
            and record["과목표현"] == subject
        ]
        write_hub(
            f"{subject}과외찾기", f"{subject}과외 지역 찾기",
            f"{subject}과외 주요 지역 페이지를 두 번의 이동 안에 탐색합니다.",
            [(str(record["path"]), str(record["keyword"])) for record in subject_records],
        )
    grade_records = [
        record for record in generated_records
        if record["page_type"] in ("region_grade", "region_grade_subject")
    ]
    write_hub(
        "학년찾기", "학년별 과외 찾기",
        "초등·중등·고등 학년별 과외 페이지를 탐색합니다.",
        [(str(record["path"]), str(record["keyword"])) for record in grade_records],
    )
    write_hub(
        "학교찾기", "학교별 과외 찾기",
        "중학교와 고등학교 과외 페이지를 학교급별로 탐색합니다.",
        [
            (public_path(category, "중학교찾기"), "중학교 과외 찾기"),
            (public_path(category, "고등학교찾기"), "고등학교 과외 찾기"),
        ],
    )
    for grade in ("중학교", "고등학교"):
        school_records = [
            record for record in generated_records
            if record["scope"] == "school" and record["학교급"] == grade
            and record["page_type"] == "school_base"
        ]
        write_hub(
            f"{grade}찾기", f"{grade} 과외 찾기",
            f"{grade}별 과목 과외와 내신 페이지를 탐색합니다.",
            [(str(record["path"]), str(record["keyword"])) for record in school_records],
        )

    # 지역 수가 많은 집중 검수에서도 홈→검수 허브→묶음→페이지의 3단계를 보장한다.
    batch_hub_links: list[tuple[str, str]] = []
    batch_size = max(10, int(config["internal_link_max"]) - 5)
    for offset in range(0, len(generated_records), batch_size):
        number = offset // batch_size + 1
        slug = f"전체샘플찾기{number}"
        batch_hub_links.append((public_path(category, slug), f"검수 페이지 묶음 {number}"))
        write_hub(
            slug, f"검수 페이지 묶음 {number}",
            "집중 품질검수 페이지를 실제 내부 링크로 연결합니다.",
            [
                (str(record["path"]), str(record["keyword"]))
                for record in generated_records[offset:offset + batch_size]
            ],
        )
    # 홈의 30링크 제한을 지키면서 모든 묶음을 연결하기 위한 결정적 그룹 계층이다.
    batch_group_links: list[tuple[str, str]] = []
    batch_group_size = max(
        10, int(config["internal_link_max"]) - len(common_hub_links) - 1
    )
    for offset in range(0, len(batch_hub_links), batch_group_size):
        number = offset // batch_group_size + 1
        slug = f"전체샘플그룹{number}"
        batch_group_links.append(
            (public_path(category, slug), f"검수 페이지 그룹 {number}")
        )
        write_hub(
            slug, f"검수 페이지 그룹 {number}",
            "관련 검수 페이지 묶음을 실제 내부 링크로 연결합니다.",
            batch_hub_links[offset:offset + batch_group_size],
        )
    write_hub(
        "전체샘플찾기", "전체 검수 페이지 찾기",
        "품질검수 페이지를 그룹과 묶음 순서로 찾아 확인합니다.",
        batch_group_links,
    )

    home_links = common_hub_links + batch_group_links + [
        (str(record["path"]), str(record["keyword"])) for record in region_base_records
    ] + [
        (str(record["path"]), str(record["keyword"])) for record in generated_records
        if record["page_type"] == "metro_school_directory"
    ]
    home_links = list(dict.fromkeys(home_links))[: int(config["internal_link_max"])]
    home_html = render(home_template, {
        "language": html.escape(str(settings["language"]), quote=True),
        "title": html.escape(f"{settings['site_name']} 지역·학교 과외 탐색"),
        "description": "지역, 과목, 학년, 학교별 과외 페이지를 빠르게 탐색할 수 있습니다.",
        "canonical": html.escape(site_url + "/", quote=True),
        "heading": html.escape(str(settings["site_name"])),
        "body_html": "지역 대표 페이지와 과목·학년·학교 탐색 페이지를 이용하세요.",
        "internal_links": "\n".join(
            f'      <a href="{href}">{html.escape(label)}</a>'
            for href, label in home_links
        ),
    })
    (output / "index.html").write_text(home_html, encoding="utf-8")

    urls_by_family: dict[str, list[str]] = defaultdict(list)
    urls_by_family["navigation"] = [site_url + "/"] + [
        page["canonical"] for page in navigation_pages
    ]
    for record in generated_records:
        if record["scope"] == "school":
            family = "school"
        elif record["scope"] == "directory":
            family = "directory"
        elif record["page_type"] in ("region_subject", "region_subject_exam"):
            family = "subject"
        elif record["page_type"] in (
            "region_grade", "region_grade_subject", "region_grade_exam",
            "region_grade_subject_exam",
        ):
            family = "grade"
        else:
            family = "region"
        urls_by_family[family].append(str(record["canonical"]))
    urls = [url for family_urls in urls_by_family.values() for url in family_urls]
    sitemap_limit = int(config["sitemap_max_urls"])
    sitemap_files: list[str] = []
    for family in ("navigation", "region", "subject", "grade", "school", "directory"):
        family_urls = urls_by_family.get(family, [])
        for index in range(0, len(family_urls), sitemap_limit):
            number = index // sitemap_limit + 1
            name = f"sitemap-{family}-{number:03d}.xml"
            lines = [
                '<?xml version="1.0" encoding="UTF-8"?>',
                '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
                *[
                    f"  <url><loc>{xml_escape(url)}</loc></url>"
                    for url in family_urls[index:index + sitemap_limit]
                ],
                "</urlset>",
            ]
            (output / name).write_text("\n".join(lines) + "\n", encoding="utf-8")
            sitemap_files.append(name)
    sitemap_index_lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
        *[
            f"  <sitemap><loc>{xml_escape(site_join(site_url, '/' + name))}</loc></sitemap>"
            for name in sitemap_files
        ],
        "</sitemapindex>",
    ]
    (output / "sitemap_index.xml").write_text(
        "\n".join(sitemap_index_lines) + "\n", encoding="utf-8"
    )
    (output / "robots.txt").write_text(
        "User-agent: *\nAllow: /\n\n"
        f"Sitemap: {site_join(site_url, '/sitemap_index.xml')}\n",
        encoding="utf-8",
    )
    return {
        "records": generated_records,
        "navigation_pages": navigation_pages,
        "urls_by_family": dict(urls_by_family),
        "generated_content_page_count": len(generated_records),
        "html_count": len(generated_records) + len(navigation_pages) + 1,
        "sitemap_url_count": len(urls),
        "copied_image_count": len(copied),
        "title_retry_count": retry_count,
        "description_duplicate_count": len(generated_records) - len(used_descriptions),
        "body_duplicate_count": len(generated_records) - len(used_bodies),
        "duration_seconds": round(time.perf_counter() - started, 3),
    }


def validate_preview(
    output: Path,
    generated: dict[str, object],
    schools: dict[str, dict[str, object]],
    settings: dict[str, object],
) -> dict[str, object]:
    records = list(generated["records"])
    slugs = [str(record["slug"]) for record in records]
    canonicals = [str(record["canonical"]) for record in records]
    title_missing = sum(str(record["keyword"]) not in str(record["title"]) for record in records)
    description_missing = sum(str(record["keyword"]) not in str(record["description"]) for record in records)
    body_missing = sum(str(record["keyword"]) not in str(record["body"]) for record in records)
    elementary_exam = sum(
        str(record["학년표현"]).startswith("초등") and bool(record["내신사용"])
        for record in records
    )
    school_bad_keyword = 0
    school_bad_title = 0
    nonexistent_school = 0
    for record in records:
        if record["scope"] != "school":
            continue
        school = str(record["학교명"])
        keyword = str(record["keyword"])
        allowed = {school + suffix for suffix in SCHOOL_ALLOWED_SUFFIXES}
        school_bad_keyword += int(keyword not in allowed)
        school_bad_title += int(any(
            school + token in str(record["title"]) for token in SCHOOL_FORBIDDEN_AFTER_NAME
        ))
        nonexistent_school += int(school not in schools)

    broken_links: list[str] = []
    broken_images: list[str] = []
    invalid_image_attribute_count = 0
    missing_stylesheet_reference_count = 0
    inline_style_block_count = 0
    script_tag_count = 0
    external_font_request_count = 0
    incoming = Counter()
    href_re = re.compile(r'<a[^>]+href=["\']([^"\']+)["\']', re.IGNORECASE)
    image_re = re.compile(r'<img[^>]+src=["\']([^"\']+)["\']', re.IGNORECASE)
    image_tag_re = re.compile(r"<img\b[^>]*>", re.IGNORECASE)
    html_files = list(output.rglob("*.html"))
    page_paths: dict[str, dict[str, object]] = {}
    file_url_paths: dict[Path, str] = {}
    for html_file in html_files:
        relative = html_file.relative_to(output)
        if relative == Path("index.html"):
            url = "/"
        else:
            url = "/" + "/".join(
                quote(part, safe="") for part in relative.parent.parts
            ) + "/"
        file_url_paths[html_file] = url
        page_paths[url] = {}
    graph: dict[str, set[str]] = defaultdict(set)
    internal_link_counts: dict[str, int] = {}
    html_sizes: dict[str, int] = {}
    for html_file in html_files:
        text = html_file.read_text(encoding="utf-8")
        missing_stylesheet_reference_count += int(
            'href="/assets/site.css"' not in text
        )
        inline_style_block_count += len(re.findall(r"<style\b", text, re.IGNORECASE))
        script_tag_count += len(re.findall(r"<script\b", text, re.IGNORECASE))
        external_font_request_count += len(
            re.findall(r"(?:fonts\.googleapis|fonts\.gstatic|@font-face)", text, re.IGNORECASE)
        )
        source_url = file_url_paths[html_file]
        html_sizes[source_url] = html_file.stat().st_size
        valid_page_links: set[str] = set()
        for href in href_re.findall(text):
            parsed = urlsplit(html.unescape(href))
            if parsed.scheme or parsed.netloc:
                continue
            path = parsed.path
            if path not in page_paths:
                broken_links.append(f"{html_file}: {href}")
            if path in page_paths and path != source_url:
                incoming[path] += 1
                valid_page_links.add(path)
                graph[source_url].add(path)
        internal_link_counts[source_url] = len(valid_page_links)
        for src in image_re.findall(text):
            parsed = urlsplit(html.unescape(src))
            target = output / Path(unquote(parsed.path).lstrip("/").replace("/", "\\"))
            if not target.is_file():
                broken_images.append(f"{html_file}: {src}")
        for image_tag in image_tag_re.findall(text):
            invalid_image_attribute_count += int(
                not all(
                    re.search(rf'\b{name}=["\'][^"\']+["\']', image_tag, re.IGNORECASE)
                    for name in ("width", "height", "loading", "decoding")
                )
            )
    orphan_pages = [
        path for path in page_paths
        if path != "/" and incoming[path] == 0
    ]
    minimum_links = int(settings["school_region_generation"]["internal_link_min"])
    maximum_links = int(settings["school_region_generation"]["internal_link_max"])
    internal_link_under_minimum = sum(
        count < minimum_links for count in internal_link_counts.values()
    )
    internal_link_over_maximum = sum(
        count > maximum_links for count in internal_link_counts.values()
    )
    depths = {"/": 0}
    queue = deque(["/"])
    while queue:
        current = queue.popleft()
        for target in graph.get(current, set()):
            if target not in depths:
                depths[target] = depths[current] + 1
                queue.append(target)
    unreachable_from_home = [path for path in page_paths if path not in depths]
    max_click_depth = max(depths.values(), default=0)
    record_by_path = {
        str(record["path"]): record for record in records
    }
    important_depth_errors = [
        path for path, record in record_by_path.items()
        if record["page_type"] in ("region_base", "region_subject")
        and depths.get(path, 99) > 2
    ]
    over_three_depth = [
        path for path in page_paths if depths.get(path, 99) > 3
    ]
    html_count = len(html_files)
    sitemap_url_count = sum(
        len(re.findall(r"<loc>.*?</loc>", path.read_text(encoding="utf-8"), re.DOTALL))
        for path in output.glob("sitemap-*.xml")
    )
    result = {
        "generated_html_count": html_count,
        "expected_html_count": int(generated["html_count"]),
        "sitemap_url_count": sitemap_url_count,
        "expected_sitemap_url_count": int(generated["sitemap_url_count"]),
        "duplicate_slug_count": len(slugs) - len(set(slugs)),
        "duplicate_canonical_count": len(canonicals) - len(set(canonicals)),
        "broken_internal_link_count": len(broken_links),
        "broken_internal_links": broken_links[:100],
        "broken_image_count": len(broken_images),
        "broken_images": broken_images[:100],
        "orphan_page_count": len(orphan_pages),
        "css_file_count": len(list(output.rglob("*.css"))),
        "missing_stylesheet_reference_count": missing_stylesheet_reference_count,
        "inline_style_block_count": inline_style_block_count,
        "script_tag_count": script_tag_count,
        "external_font_request_count": external_font_request_count,
        "invalid_image_attribute_count": invalid_image_attribute_count,
        "orphan_pages": orphan_pages[:100],
        "pages_below_internal_link_minimum": internal_link_under_minimum,
        "pages_above_internal_link_maximum": internal_link_over_maximum,
        "internal_link_minimum": min(internal_link_counts.values(), default=0),
        "internal_link_maximum": max(internal_link_counts.values(), default=0),
        "home_max_click_depth": max_click_depth,
        "pages_over_three_clicks": len(over_three_depth),
        "important_pages_over_two_clicks": len(important_depth_errors),
        "unreachable_from_home_count": len(unreachable_from_home),
        "title_keyword_missing_count": title_missing,
        "description_keyword_missing_count": description_missing,
        "body_keyword_missing_count": body_missing,
        "elementary_internal_exam_count": elementary_exam,
        "school_invalid_keyword_count": school_bad_keyword,
        "school_unnatural_grade_title_count": school_bad_title,
        "nonexistent_school_count": nonexistent_school,
        "description_duplicate_count": int(generated["description_duplicate_count"]),
        "body_duplicate_count": int(generated["body_duplicate_count"]),
        "empty_page_count": sum(not str(record["body"]).strip() for record in records),
        "title_over_max_count": sum(
            len(str(record["title"])) > int(settings["content_generation"]["title_max_length"])
            for record in records
        ),
        "html_size_min_bytes": min(html_sizes.values(), default=0),
        "html_size_max_bytes": max(html_sizes.values(), default=0),
        "html_size_average_bytes": round(
            sum(html_sizes.values()) / max(1, len(html_sizes)), 2
        ),
        "html_over_hard_max_count": sum(
            size > int(settings["school_region_generation"]["html_hard_max_bytes"])
            for size in html_sizes.values()
        ),
        "html_below_target_count": sum(
            size < int(settings["school_region_generation"]["html_target_min_bytes"])
            for size in html_sizes.values()
        ),
        "page_metrics": [
            {
                "url_path": path,
                "click_depth": depths.get(path),
                "internal_link_count": internal_link_counts.get(path, 0),
                "html_size_bytes": html_sizes.get(path, 0),
                "is_orphan": path in orphan_pages,
            }
            for path in sorted(page_paths)
        ],
    }
    zero_required = (
        "duplicate_slug_count", "duplicate_canonical_count",
        "broken_internal_link_count", "broken_image_count", "orphan_page_count",
        "missing_stylesheet_reference_count", "inline_style_block_count",
        "script_tag_count", "external_font_request_count",
        "invalid_image_attribute_count",
        "pages_below_internal_link_minimum", "pages_above_internal_link_maximum",
        "pages_over_three_clicks", "important_pages_over_two_clicks",
        "unreachable_from_home_count", "title_keyword_missing_count",
        "description_keyword_missing_count", "body_keyword_missing_count",
        "elementary_internal_exam_count", "school_invalid_keyword_count",
        "school_unnatural_grade_title_count", "nonexistent_school_count",
        "description_duplicate_count", "body_duplicate_count", "empty_page_count",
        "title_over_max_count", "html_over_hard_max_count",
    )
    result["errors"] = [
        f"{key}={result[key]}" for key in zero_required if result[key] != 0
    ]
    if html_count != int(generated["html_count"]):
        result["errors"].append("HTML 수 불일치")
    if sitemap_url_count != int(generated["sitemap_url_count"]):
        result["errors"].append("sitemap URL 수 불일치")
    if result["css_file_count"] != 1:
        result["errors"].append(f"공통 CSS 파일 수 오류: {result['css_file_count']}")
    result["passed"] = not result["errors"]
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="지역·학교 계층형 페이지 계획 및 후보 생성")
    parser.add_argument("--input", required=True, help="원본 Excel 경로")
    args = parser.parse_args(argv)
    input_path = Path(args.input).resolve()
    if not input_path.is_file():
        print(f"입력 Excel을 찾을 수 없습니다: {input_path}", file=sys.stderr)
        return 1
    settings = json.loads(SETTINGS_PATH.read_text(encoding="utf-8-sig"))
    config = dict(settings["school_region_generation"])
    links, regions, unmatched, normalization = read_excel(input_path)
    schools = build_school_records(links)
    plans, duplicate_rows, school_validation = plan_pages(
        regions, links, schools, config
    )
    enabled_phases = set(str(value) for value in config["enabled_generation_phases"])
    plans = [
        plan for plan in plans if generation_phase(plan) in enabled_phases
    ]
    sample, selected_regions, selected_schools = select_sample(
        plans, regions, schools, config
    )

    REPORTS.mkdir(parents=True, exist_ok=True)
    type_counts = Counter(str(plan["page_type"]) for plan in plans)
    invalid_rows = [
        {
            "오류유형": "금지 조합",
            "page_type": plan["page_type"],
            "지역명": plan["지역명"],
            "학교명": plan["학교명"],
            "keyword": plan["keyword"],
            "validation_note": "초등 내신 또는 학교급 중복",
        }
        for plan in plans
        if (
            str(plan["학년표현"]).startswith("초등") and bool(plan["내신사용"])
        ) or (
            plan["scope"] == "school"
            and any(
                str(plan["학교명"]) + token in str(plan["keyword"])
                for token in SCHOOL_FORBIDDEN_AFTER_NAME
            )
        )
    ]
    representative_rows = [
        {
            "학교명": name,
            "학교급": school["학교급"],
            "대표지역": school["대표지역"],
            "시도": school["시도"],
            "행정구": school["행정구"],
            "매칭기준": school["매칭기준"],
            "관련지역수": len(school["관련지역"]),
            "관련지역": "|".join(school["관련지역"]),
            "서로다른주소수": school["서로다른주소수"],
        }
        for name, school in sorted(schools.items())
    ]
    write_csv(
        REPORTS / "page_plan_by_type.csv",
        ("page_type", "expected_count"),
        [{"page_type": key, "expected_count": value} for key, value in sorted(type_counts.items())],
    )
    write_plan_xlsx(REPORTS / "page_plan_sample.xlsx", sample)
    write_csv(
        REPORTS / "duplicate_slug_report.csv",
        ("slug", "page_type", "지역명", "학교명", "처리"),
        duplicate_rows,
    )
    write_csv(
        REPORTS / "invalid_combination_report.csv",
        ("오류유형", "page_type", "지역명", "학교명", "keyword", "validation_note"),
        invalid_rows,
    )
    write_csv(
        REPORTS / "school_representative_region_report.csv",
        ("학교명", "학교급", "대표지역", "시도", "행정구", "매칭기준", "관련지역수", "관련지역", "서로다른주소수"),
        representative_rows,
    )
    write_csv(
        REPORTS / "unmatched_region_report.csv",
        ("지역명", "상태"),
        unmatched,
    )
    write_csv(
        REPORTS / "school_keyword_validation_report.csv",
        ("학교명", "학교급", "keyword", "slug", "허용조합", "학교급중복표현", "validation_note"),
        school_validation,
    )

    output = choose_preview_dir()
    generated = generate_preview(
        output, sample, selected_regions, schools, links, settings, config
    )
    validation = validate_preview(output, generated, schools, settings)
    write_csv(
        REPORTS / "page_html_size_report.csv",
        ("url_path", "click_depth", "internal_link_count", "html_size_bytes", "is_orphan"),
        list(validation["page_metrics"]),
    )
    output_size = sum(path.stat().st_size for path in output.rglob("*") if path.is_file())
    navigation_page_count = 12
    total_pages_with_navigation = len(plans) + navigation_page_count
    sample_html_bytes = sum(
        path.stat().st_size for path in output.rglob("*.html")
    )
    shared_asset_bytes = output_size - sample_html_bytes
    average_html_bytes = sample_html_bytes / max(1, int(generated["html_count"]))
    estimated_size = round(
        shared_asset_bytes
        + average_html_bytes * total_pages_with_navigation
        + len(plans) * 220
    )
    estimated_time = round(
        float(generated["duration_seconds"]) / max(1, int(generated["generated_content_page_count"]))
        * len(plans), 1
    )
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_excel": str(input_path),
        "input_sheets": ["지역별 학교", "지역 요약", "미매칭 지역"],
        "normalization": normalization,
        "region_count": len(regions),
        "matched_region_count": len({str(row["지역명"]) for row in links}),
        "unmatched_region_count": len(unmatched),
        "unique_middle_school_count": sum(
            school["학교급"] == "중학교" for school in schools.values()
        ),
        "unique_high_school_count": sum(
            school["학교급"] == "고등학교" for school in schools.values()
        ),
        "unique_school_name_count": len(schools),
        "page_type_counts": dict(sorted(type_counts.items())),
        "expected_total_content_page_count": len(plans),
        "expected_total_page_count_with_navigation": total_pages_with_navigation,
        "generation_order": config["generation_order"],
        "enabled_generation_phases": config["enabled_generation_phases"],
        "duplicate_slug_count": len(duplicate_rows),
        "invalid_combination_count": len(invalid_rows),
        "school_invalid_combination_count": sum(
            bool(row["학교급중복표현"]) or not bool(row["허용조합"])
            for row in school_validation
        ),
        "selected_sample_regions": selected_regions,
        "selected_sample_schools": selected_schools,
        "sample_content_page_count": int(generated["generated_content_page_count"]),
        "sample_html_count": int(generated["html_count"]),
        "sample_sitemap_url_count": int(generated["sitemap_url_count"]),
        "preview_output": str(output),
        "sample_output_size_bytes": output_size,
        "sample_generation_seconds": generated["duration_seconds"],
        "estimated_full_output_size_bytes": estimated_size,
        "estimated_full_generation_seconds": estimated_time,
        "validation": validation,
        "representative_preview_paths": [
            str(record["path"]) for record in list(generated["records"])[:10]
        ],
        "generated_reports": [
            "reports/page_plan_summary.json",
            "reports/page_plan_by_type.csv",
            "reports/page_plan_sample.xlsx",
            "reports/duplicate_slug_report.csv",
            "reports/invalid_combination_report.csv",
            "reports/school_representative_region_report.csv",
            "reports/unmatched_region_report.csv",
            "reports/school_keyword_validation_report.csv",
            "reports/school_region_preview_validation.json",
            "reports/page_html_size_report.csv",
        ],
    }
    (REPORTS / "page_plan_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (REPORTS / "school_region_preview_validation.json").write_text(
        json.dumps(validation, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if validation["passed"] and not invalid_rows else 1


if __name__ == "__main__":
    raise SystemExit(main())
