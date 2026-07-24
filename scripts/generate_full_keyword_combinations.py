from __future__ import annotations

import csv
import hashlib
import heapq
import json
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from keyword_combination_engine import (
    EXAM_GRADES,
    GRADES,
    SUBJECTS,
    core_terms,
    display_keyword,
    make_title,
    pick,
    stable_int,
)


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
REPORTS = ROOT / "reports"
REGION_SOURCE = DATA / "gongbuup_region_master.csv"
SCHOOL_SOURCE = DATA / "gongbuup_school_master.csv"
PROFILE_SOURCE = DATA / "keyword_profiles" / "keyword_combinations.xlsx"
SETTINGS_SOURCE = ROOT / "config" / "settings.json"
FULL_CSV = DATA / "full_keyword_combinations.csv"
FULL_XLSX = DATA / "full_keyword_combinations.xlsx"
CONFLICT_CSV = DATA / "full_keyword_conflicts.csv"
INVALID_CSV = DATA / "full_keyword_invalid.csv"
SAMPLE_CSV = DATA / "full_validation_sample_1000.csv"
REPORT_JSON = REPORTS / "full_keyword_validation_report.json"
REPORT_TXT = REPORTS / "full_keyword_validation_report.txt"

FULL_FIELDS = [
    "page_type", "시도", "시군구", "읍면동", "리", "지역명", "학교명", "학교급",
    "학년표현", "과목표현", "내신사용", "display_keyword", "compact_keyword",
    "slug", "title", "description", "title_core_terms", "content_seed",
    "is_valid", "validation_note",
]
CONFLICT_FIELDS = [
    "오류 또는 충돌 유형", "page_type", "지역명", "학교명", "기존 slug", "최종 slug",
    "title", "충돌 대상", "해결 여부", "validation_note",
]
WINDOWS_INVALID_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]|[. ]$')
ABNORMAL_RE = re.compile(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff\ufffd]")
PROVINCE_SHORT = {
    "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
    "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
    "울산광역시": "울산", "세종특별자치시": "세종", "경기도": "경기",
    "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남",
    "전북특별자치도": "전북", "전라남도": "전남", "경상북도": "경북",
    "경상남도": "경남", "제주특별자치도": "제주",
}


def read_profiles() -> list[dict[str, str]]:
    workbook = load_workbook(PROFILE_SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(rows)]
    result = [
        {header: str(value or "") for header, value in zip(headers, values)}
        for values in rows
    ]
    workbook.close()
    return [
        row for row in result
        if row["활성화"].lower() == "true" and row["page_scope"] in ("region", "school", "both")
    ]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [row for row in csv.DictReader(handle) if row.get("is_valid", "").lower() == "true"]


def profile_combo(profile: dict[str, str]) -> dict[str, object]:
    return {
        "type": profile["비고"],
        "grade": profile["학년표현"],
        "subject": profile["과목표현"],
        "exam": profile["내신사용"].lower() == "true",
        "id": profile["combination_id"],
    }


def compatible_school(profile: dict[str, str], school_grade: str) -> bool:
    if profile["page_scope"] not in ("school", "both"):
        return False
    grade = profile["학년표현"]
    if school_grade == "중학교":
        return grade in ("중등", "중학생")
    if school_grade == "고등학교":
        return grade in ("고등", "고등학생")
    return False


def create_description(
    keyword: str, seed: str, description_pool: list[dict[str, str]]
) -> str:
    used: set[str] = set()
    items = [pick(description_pool, seed, f"description_{index}", used) for index in range(4)]
    description = f"{keyword}를 살펴볼 때에는 {items[0]['text']} " + " ".join(item["text"] for item in items[1:])
    if len(description) < 250:
        description += " 학습한 내용과 실제 수행 결과를 비교하면 다음 공부의 우선순위를 더 구체적으로 정할 수 있습니다."
    return re.sub(r"\s+", " ", description).strip()


def choose_slug(
    base: str,
    row: dict[str, str],
    used: dict[str, str],
) -> tuple[str, str, int]:
    if base not in used:
        return base, "", 0
    district = re.sub(r"\s+", "", row["시군구"])
    candidate = f"{district}{base}"
    if candidate not in used:
        return candidate, "시군구 보강", 1
    province = PROVINCE_SHORT.get(row["시도"], re.sub(r"\s+", "", row["시도"]))
    candidate = f"{province}{district}{base}"
    if candidate not in used:
        return candidate, "시도축약·시군구 보강", 1
    suffix = 2
    while f"{candidate}{suffix}" in used:
        suffix += 1
    return f"{candidate}{suffix}", "숫자 suffix 최종 보강", 1


def xlsx_append_setup(fields: list[str]) -> tuple[Workbook, object, list[int]]:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("전체")
    sheet.freeze_panes = "A2"
    sheet.append(fields)
    return workbook, sheet, [len(field) for field in fields]


def update_widths(widths: list[int], values: list[object]) -> None:
    for index, value in enumerate(values):
        widths[index] = min(60, max(widths[index], len(str(value or ""))))


def finish_xlsx(workbook: Workbook, sheet: object, widths: list[int], row_count: int) -> None:
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{row_count + 1}"
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width + 3, 60)
    workbook.save(FULL_XLSX)
    # In openpyxl write-only mode, column dimensions assigned after sheet data
    # are not serialized because <cols> must precede <sheetData>. Inject the
    # measured widths into the worksheet XML without loading the large workbook.
    cols = ["<cols>"]
    for index, width in enumerate(widths, start=1):
        cols.append(
            f'<col min="{index}" max="{index}" width="{min(width + 3, 60)}" '
            'customWidth="1"/>'
        )
    cols.append("</cols>")
    cols_xml = "".join(cols).encode("utf-8")
    temporary = FULL_XLSX.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(FULL_XLSX, "r") as source, zipfile.ZipFile(
        temporary, "w"
    ) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b"<sheetData>", cols_xml + b"<sheetData>", 1)
            target.writestr(info, payload)
    temporary.replace(FULL_XLSX)


def add_sample_candidate(
    heaps: dict[tuple[str, ...], list[tuple[int, str, dict[str, str]]]],
    key: tuple[str, ...],
    row: dict[str, str],
    capacity: int,
) -> None:
    score = stable_int(f"{row['content_seed']}:sample")
    entry = (-score, row["slug"], row)
    heap = heaps[key]
    if len(heap) < capacity:
        heapq.heappush(heap, entry)
    elif entry > heap[0]:
        heapq.heapreplace(heap, entry)


def balanced_from_heaps(
    heaps: dict[tuple[str, ...], list[tuple[int, str, dict[str, str]]]],
    count: int,
) -> list[dict[str, str]]:
    groups = {
        key: [entry[2] for entry in sorted(values, key=lambda item: (-item[0], item[1]))]
        for key, values in heaps.items()
    }
    keys = sorted(groups)
    result: list[dict[str, str]] = []
    positions = Counter()
    while len(result) < count:
        progressed = False
        for key in keys:
            if positions[key] < len(groups[key]) and len(result) < count:
                result.append(groups[key][positions[key]])
                positions[key] += 1
                progressed = True
        if not progressed:
            raise ValueError(f"검수 샘플 {count}건을 선택할 수 없습니다.")
    return result


def main() -> int:
    targets = [FULL_CSV, FULL_XLSX, CONFLICT_CSV, INVALID_CSV, SAMPLE_CSV, REPORT_JSON, REPORT_TXT]
    existing = [str(path) for path in targets if path.exists()]
    if existing:
        raise FileExistsError("기존 출력 파일을 덮어쓰지 않습니다:\n" + "\n".join(existing))

    settings = json.loads(SETTINGS_SOURCE.read_text(encoding="utf-8-sig"))
    title_min = int(settings["content_generation"]["title_min_length"])
    title_max = int(settings["content_generation"]["title_max_length"])
    profiles = read_profiles()
    region_profiles = [row for row in profiles if row["page_scope"] in ("region", "both")]
    school_profiles = [row for row in profiles if row["page_scope"] in ("school", "both")]
    regions = read_csv(REGION_SOURCE)
    schools = read_csv(SCHOOL_SOURCE)
    description_pool = json.loads((DATA / "sentence_pools" / "description_general.json").read_text(encoding="utf-8-sig"))
    title_patterns = json.loads((DATA / "sentence_pools" / "title_patterns.json").read_text(encoding="utf-8-sig"))
    title_modifiers = json.loads((DATA / "sentence_pools" / "title_modifiers.json").read_text(encoding="utf-8-sig"))
    title_endings = json.loads((DATA / "sentence_pools" / "title_endings.json").read_text(encoding="utf-8-sig"))

    conflict_handle = CONFLICT_CSV.open("w", encoding="utf-8-sig", newline="")
    conflict_writer = csv.DictWriter(conflict_handle, fieldnames=CONFLICT_FIELDS)
    conflict_writer.writeheader()
    invalid_handle = INVALID_CSV.open("w", encoding="utf-8-sig", newline="")
    invalid_writer = csv.DictWriter(invalid_handle, fieldnames=CONFLICT_FIELDS)
    invalid_writer.writeheader()
    full_handle = FULL_CSV.open("w", encoding="utf-8-sig", newline="")
    full_writer = csv.DictWriter(full_handle, fieldnames=FULL_FIELDS)
    full_writer.writeheader()
    workbook, sheet, widths = xlsx_append_setup(FULL_FIELDS)

    used_slugs: dict[str, str] = {}
    used_encoded_slugs: dict[str, str] = {}
    used_titles: dict[str, str] = {}
    display_seen: Counter[str] = Counter()
    compact_seen: Counter[str] = Counter()
    normalized_seen: Counter[str] = Counter()
    region_name_paths: dict[str, set[str]] = defaultdict(set)
    school_name_paths: dict[str, set[str]] = defaultdict(set)
    counters = Counter()
    title_lengths: list[int] = []
    region_sample_heaps: dict[tuple[str, ...], list[tuple[int, str, dict[str, str]]]] = defaultdict(list)
    school_sample_heaps: dict[tuple[str, ...], list[tuple[int, str, dict[str, str]]]] = defaultdict(list)

    def process(entity_row: dict[str, str], profile: dict[str, str], page_type: str) -> None:
        combo = profile_combo(profile)
        entity = entity_row["지역표시명"] if page_type == "region" else entity_row["학교명"]
        keyword = display_keyword(entity, combo)
        compact = unicodedata.normalize("NFC", re.sub(r"\s+", "", keyword))
        content_seed = hashlib.sha256(
            f"{page_type}|{entity_row['시도']}|{entity_row['시군구']}|{entity}|{profile['combination_id']}".encode("utf-8")
        ).hexdigest()
        final_slug, slug_resolution, collided = choose_slug(compact, entity_row, used_slugs)
        counters["base_slug_collision"] += collided
        counters["administrative_resolution"] += int(collided and "보강" in slug_resolution)
        if collided:
            conflict_writer.writerow({
                "오류 또는 충돌 유형": "기본 slug 충돌",
                "page_type": page_type, "지역명": entity if page_type == "region" else "",
                "학교명": entity if page_type == "school" else "", "기존 slug": compact,
                "최종 slug": final_slug, "title": "", "충돌 대상": used_slugs.get(compact, ""),
                "해결 여부": "true", "validation_note": slug_resolution,
            })
        encoded = quote(final_slug, safe="")
        if encoded in used_encoded_slugs:
            counters["url_encoding_collision"] += 1
        used_encoded_slugs[encoded] = final_slug
        used_slugs[final_slug] = content_seed

        display_seen[keyword] += 1
        compact_seen[compact] += 1
        normalized_seen[unicodedata.normalize("NFKC", compact)] += 1
        if page_type == "region":
            region_name_paths[entity].add(f"{entity_row['시도']}|{entity_row['시군구']}|{entity_row['읍면동']}|{entity_row['리']}")
        else:
            school_name_paths[entity].add(f"{entity_row['시도']}|{entity_row['시군구']}")

        title, pattern_id, modifier, ending = make_title(
            final_slug, entity, keyword, combo, title_patterns, title_modifiers, title_endings
        )
        initial_title = title
        retry = 0
        # make_title() already evaluates 100 deterministic combinations for each
        # identity.  Use four ordered identity changes (pattern, modifier, ending,
        # helper) before the deterministic administrative qualifier fallback.
        # This remains within the configured maximum of 100 retries without
        # multiplying the inner 100-candidate search by another factor of 100.
        while title in used_titles and retry < 4:
            retry += 1
            title, pattern_id, modifier, ending = make_title(
                f"{final_slug}:title:retry_{retry}", entity, keyword, combo,
                title_patterns, title_modifiers, title_endings,
            )
        if title in used_titles:
            qualifiers = [
                entity_row["시군구"],
                f"{PROVINCE_SHORT.get(entity_row['시도'], entity_row['시도'])} {entity_row['시군구']}",
            ]
            for qualifier in qualifiers:
                candidates = [
                    f"{keyword} {qualifier} 기준으로 학습 방향과 오답 관리 과정을 정리합니다",
                    f"{keyword} {qualifier} 맞춤 학습 기준과 공부 순서를 살펴봅니다",
                    f"{keyword} {qualifier} 지역의 개념 이해와 복습 방향을 점검합니다",
                ]
                for candidate in candidates:
                    candidate = re.sub(r"\s+", " ", candidate).strip()
                    if title_min <= len(candidate) <= title_max and candidate not in used_titles:
                        title = candidate
                        break
                if title not in used_titles:
                    break
        if initial_title in used_titles:
            counters["title_duplicate_occurrence"] += 1
            counters["title_regenerated"] += int(retry > 0)
            conflict_writer.writerow({
                "오류 또는 충돌 유형": "title 중복",
                "page_type": page_type, "지역명": entity if page_type == "region" else "",
                "학교명": entity if page_type == "school" else "", "기존 slug": compact,
                "최종 slug": final_slug, "title": title, "충돌 대상": used_titles.get(initial_title, ""),
                "해결 여부": str(title not in used_titles).lower(),
                "validation_note": f"SHA-256 결정적 재시도 {retry}회",
            })
        unresolved_title_duplicate = title in used_titles
        counters["final_title_duplicate"] += int(unresolved_title_duplicate)
        used_titles[title] = final_slug
        counters["max_title_retry"] = max(counters["max_title_retry"], retry)

        description = create_description(keyword, content_seed, description_pool)
        terms = core_terms(entity, combo)
        notes: list[str] = []
        if WINDOWS_INVALID_RE.search(final_slug) or final_slug.upper() in {"CON", "PRN", "AUX", "NUL"}:
            notes.append("Windows 폴더명 부적합")
            counters["windows_invalid"] += 1
        if not (title_min <= len(title) <= title_max):
            notes.append("제목 길이 오류")
        if any(term not in title for term in terms):
            notes.append("제목 필수 키워드 누락")
            counters["required_term_missing"] += 1
        if unresolved_title_duplicate:
            notes.append("최종 title 중복")
        if combo["exam"] and combo["grade"] == "초등":
            notes.append("초등 + 내신 금지")
            counters["elementary_exam"] += 1
        if combo["exam"] and combo["grade"] == "초등학생":
            notes.append("초등학생 + 내신 금지")
            counters["elementary_student_exam"] += 1
        if page_type == "school" and not compatible_school(profile, entity_row["학교급"]):
            notes.append("학교급과 학년 표현 불일치")
            counters["school_grade_mismatch"] += 1
        if page_type == "region" and not entity:
            notes.append("빈 지역명")
        if page_type == "region" and re.search(r"(수학|영어|국영수|초등|중등|고등)$", entity):
            notes.append("지역명에 과목·학년 파생 표현 혼입")
        if page_type == "school" and not entity:
            notes.append("빈 학교명")
        if ABNORMAL_RE.search(" ".join((keyword, title, description))):
            notes.append("비정상 문자")
            counters["abnormal"] += 1
        valid = not notes
        if not valid:
            counters["invalid"] += 1
            invalid_writer.writerow({
                "오류 또는 충돌 유형": " | ".join(notes), "page_type": page_type,
                "지역명": entity if page_type == "region" else "",
                "학교명": entity if page_type == "school" else "",
                "기존 slug": compact, "최종 slug": final_slug, "title": title,
                "충돌 대상": "", "해결 여부": "false", "validation_note": "; ".join(notes),
            })

        full_row = {
            "page_type": page_type, "시도": entity_row["시도"], "시군구": entity_row["시군구"],
            "읍면동": entity_row.get("읍면동", ""), "리": entity_row.get("리", ""),
            "지역명": entity if page_type == "region" else "",
            "학교명": entity if page_type == "school" else "",
            "학교급": entity_row.get("학교급", ""), "학년표현": combo["grade"],
            "과목표현": combo["subject"], "내신사용": str(bool(combo["exam"])).lower(),
            "display_keyword": keyword, "compact_keyword": compact, "slug": final_slug,
            "title": title, "description": description, "title_core_terms": "|".join(terms),
            "content_seed": content_seed, "is_valid": str(valid).lower(),
            "validation_note": "; ".join(notes) or "정상",
        }
        full_writer.writerow(full_row)
        values = [full_row[field] for field in FULL_FIELDS]
        sheet.append(values)
        update_widths(widths, values)
        counters["total"] += 1
        counters[f"{page_type}_total"] += 1
        counters["valid"] += int(valid)
        title_lengths.append(len(title))
        if valid:
            if page_type == "region":
                add_sample_candidate(region_sample_heaps, ("region", entity_row["시도"]), full_row, 60)
            else:
                add_sample_candidate(
                    school_sample_heaps,
                    ("school", entity_row["시도"], entity_row["학교급"]),
                    full_row, 35,
                )
        if counters["total"] % 10000 == 0:
            print(f"{counters['total']} combinations", flush=True)

    try:
        for entity_row in sorted(regions, key=lambda row: (
            row["시도"], row["시군구"], row["읍면동"], row["리"], row["지역표시명"]
        )):
            for profile in region_profiles:
                process(entity_row, profile, "region")
        for entity_row in sorted(schools, key=lambda row: (
            row["시도"], row["시군구"], row["학교명"]
        )):
            for profile in school_profiles:
                if compatible_school(profile, entity_row["학교급"]):
                    process(entity_row, profile, "school")
    finally:
        full_handle.close()
        conflict_handle.close()
        invalid_handle.close()

    finish_xlsx(workbook, sheet, widths, counters["total"])
    region_sample = balanced_from_heaps(region_sample_heaps, 600)
    school_sample = balanced_from_heaps(school_sample_heaps, 400)
    with SAMPLE_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FULL_FIELDS)
        writer.writeheader()
        writer.writerows(region_sample + school_sample)

    display_duplicate_count = sum(count - 1 for count in display_seen.values() if count > 1)
    compact_duplicate_count = sum(count - 1 for count in compact_seen.values() if count > 1)
    normalized_duplicate_count = sum(count - 1 for count in normalized_seen.values() if count > 1)
    final_title_duplicates = counters["final_title_duplicate"]
    final_slug_duplicates = len(used_slugs) - len(set(used_slugs))
    title_below = sum(length < title_min for length in title_lengths)
    title_above = sum(length > title_max for length in title_lengths)
    sample_subjects = Counter(row["과목표현"] for row in region_sample + school_sample if row["과목표현"])
    sample_grades = Counter(row["학교급"] for row in school_sample)
    sample_exam = sum(row["내신사용"] == "true" for row in region_sample + school_sample)
    sample_elementary_exam = sum(
        row["내신사용"] == "true" and row["학년표현"] in ("초등", "초등학생")
        for row in region_sample + school_sample
    )
    expected_sitemap = counters["valid"] + 1
    report = {
        "total_combination_count": counters["total"],
        "region_expected_page_count": counters["region_total"],
        "school_expected_page_count": counters["school_total"],
        "final_valid_page_count": counters["valid"],
        "excluded_page_count": counters["invalid"],
        "base_slug_collision_count": counters["base_slug_collision"],
        "administrative_resolution_count": counters["administrative_resolution"],
        "unresolved_slug_collision_count": final_slug_duplicates,
        "display_keyword_duplicate_count": display_duplicate_count,
        "compact_keyword_duplicate_count": compact_duplicate_count,
        "normalized_keyword_duplicate_count": normalized_duplicate_count,
        "title_duplicate_occurrence_count": counters["title_duplicate_occurrence"],
        "title_regenerated_count": counters["title_regenerated"],
        "title_max_retry_count": counters["max_title_retry"],
        "final_title_duplicate_count": final_title_duplicates,
        "title_minimum_length": min(title_lengths, default=0),
        "title_maximum_length": max(title_lengths, default=0),
        "title_average_length": round(sum(title_lengths) / len(title_lengths), 2) if title_lengths else 0,
        "title_below_minimum_count": title_below,
        "title_above_maximum_count": title_above,
        "required_title_term_missing_count": counters["required_term_missing"],
        "elementary_exam_count": counters["elementary_exam"],
        "elementary_student_exam_count": counters["elementary_student_exam"],
        "school_grade_mismatch_count": counters["school_grade_mismatch"],
        "windows_path_invalid_count": counters["windows_invalid"],
        "url_encoding_collision_count": counters["url_encoding_collision"],
        "abnormal_character_count": counters["abnormal"],
        "same_region_name_different_area_count": sum(len(paths) > 1 for paths in region_name_paths.values()),
        "same_school_name_different_area_count": sum(len(paths) > 1 for paths in school_name_paths.values()),
        "expected_html_file_count": counters["valid"],
        "expected_sitemap_url_count": expected_sitemap,
        "validation_sample_count": len(region_sample) + len(school_sample),
        "validation_sample_region_count": len(region_sample),
        "validation_sample_school_count": len(school_sample),
        "validation_sample_school_grade_count": dict(sample_grades),
        "validation_sample_subject_count": dict(sample_subjects),
        "validation_sample_internal_exam_count": sample_exam,
        "validation_sample_elementary_exam_count": sample_elementary_exam,
        "generated_files": [
            str(path.relative_to(ROOT)) for path in (
                FULL_CSV, FULL_XLSX, CONFLICT_CSV, INVALID_CSV, SAMPLE_CSV, REPORT_JSON, REPORT_TXT
            )
        ],
        "modified_files": ["scripts/generate_full_keyword_combinations.py"],
    }
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = [
        "전체 키워드 조합 검증 보고서",
        f"전체 조합 수: {report['total_combination_count']}",
        f"지역 페이지 예정 수: {report['region_expected_page_count']}",
        f"학교 페이지 예정 수: {report['school_expected_page_count']}",
        f"최종 유효 페이지 수: {report['final_valid_page_count']}",
        f"제외 페이지 수: {report['excluded_page_count']}",
        f"기본 slug 충돌 수: {report['base_slug_collision_count']}",
        f"행정구역 보강 후 해결 수: {report['administrative_resolution_count']}",
        f"최종 미해결 slug 충돌 수: {report['unresolved_slug_collision_count']}",
        f"display_keyword 중복 수: {report['display_keyword_duplicate_count']}",
        f"compact_keyword 중복 수: {report['compact_keyword_duplicate_count']}",
        f"title 중복 발생 수: {report['title_duplicate_occurrence_count']}",
        f"title 재생성 수: {report['title_regenerated_count']}",
        f"최종 title 중복 수: {report['final_title_duplicate_count']}",
        f"제목 최소 길이: {report['title_minimum_length']}",
        f"제목 최대 길이: {report['title_maximum_length']}",
        f"제목 평균 길이: {report['title_average_length']}",
        f"40자 미만 수: {report['title_below_minimum_count']}",
        f"68자 초과 수: {report['title_above_maximum_count']}",
        f"필수 키워드 누락 수: {report['required_title_term_missing_count']}",
        f"초등 + 내신 수: {report['elementary_exam_count']}",
        f"초등학생 + 내신 수: {report['elementary_student_exam_count']}",
        f"학교급 불일치 수: {report['school_grade_mismatch_count']}",
        f"Windows 경로 부적합 수: {report['windows_path_invalid_count']}",
        f"비정상 문자 수: {report['abnormal_character_count']}",
        f"전체 예상 HTML 파일 수: {report['expected_html_file_count']}",
        f"예상 sitemap URL 수: {report['expected_sitemap_url_count']}",
        "",
        "생성한 파일:",
        *[f"- {path}" for path in report["generated_files"]],
    ]
    REPORT_TXT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
