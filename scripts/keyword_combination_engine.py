from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
POOLS = DATA / "sentence_pools"
PROFILE_DIR = DATA / "keyword_profiles"
PROFILE_CSV = PROFILE_DIR / "keyword_combinations.csv"
PROFILE_XLSX = PROFILE_DIR / "keyword_combinations.xlsx"
SAMPLE_CSV = DATA / "keyword_combination_sample_200.csv"
SAMPLE_XLSX = DATA / "keyword_combination_sample_200.xlsx"
BUILD_INPUT = DATA / "keyword_combination_build_200.csv"
REPORT_JSON = ROOT / "reports" / "keyword_combination_report.json"
REPORT_TXT = ROOT / "reports" / "keyword_combination_report.txt"

GRADES = ["초등", "초등학생", "중등", "중학생", "고등", "고등학생"]
SUBJECTS = ["영어", "수학", "과학", "국어", "영수"]
EXAM_GRADES = ["중등", "중학생", "고등", "고등학생"]
TARGET_BY_EXPRESSION = {
    "초등": "초등학생", "초등학생": "초등학생",
    "중등": "중학생", "중학생": "중학생",
    "고등": "고등학생", "고등학생": "고등학생",
}
GRADE_POOL = {
    "초등": "body_elementary", "초등학생": "body_elementary",
    "중등": "body_middle", "중학생": "body_middle",
    "고등": "body_high", "고등학생": "body_high",
}
SUBJECT_POOL = {
    "영어": "body_english", "수학": "body_math", "과학": "body_science",
    "국어": "body_korean", "영수": "body_english_math",
}
PROFILE_FIELDS = [
    "combination_id", "page_scope", "학년표현", "과목표현", "내신사용", "키워드패턴",
    "제목사용", "본문사용", "description사용", "활성화", "가중치", "비고",
]
SAMPLE_FIELDS = [
    "page_type", "지역명", "학교명", "학교급", "학년표현", "과목표현", "내신사용",
    "display_keyword", "compact_keyword", "slug", "title", "title_core_terms",
    "body_core_terms", "description_core_terms", "is_valid", "validation_note",
]


def stable_int(value: str) -> int:
    return int(hashlib.sha256(value.encode("utf-8")).hexdigest(), 16)


def load_json(name: str) -> list[dict[str, str]]:
    return json.loads((POOLS / f"{name}.json").read_text(encoding="utf-8-sig"))


def logical_combinations() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    rows.append({"type": "base", "grade": "", "subject": "", "exam": False})
    rows.extend({"type": "subject_only", "grade": "", "subject": subject, "exam": False} for subject in SUBJECTS)
    rows.extend({"type": "grade_only", "grade": grade, "subject": "", "exam": False} for grade in GRADES)
    rows.extend(
        {"type": "grade_subject", "grade": grade, "subject": subject, "exam": False}
        for grade in GRADES for subject in SUBJECTS
    )
    rows.append({"type": "exam_only", "grade": "", "subject": "", "exam": True})
    rows.extend({"type": "subject_exam", "grade": "", "subject": subject, "exam": True} for subject in SUBJECTS)
    rows.extend({"type": "grade_exam", "grade": grade, "subject": "", "exam": True} for grade in EXAM_GRADES)
    rows.extend(
        {"type": "grade_subject_exam", "grade": grade, "subject": subject, "exam": True}
        for grade in EXAM_GRADES for subject in SUBJECTS
    )
    return rows


def keyword_pattern(scope: str, combo: dict[str, object]) -> str:
    entity = "{지역명}" if scope == "region" else "{학교명}"
    grade = "{학년}" if combo["grade"] else ""
    subject = "{과목}" if combo["subject"] else ""
    exam = "내신" if combo["exam"] else ""
    return f"{entity}{grade}{subject}{exam}과외"


def create_profiles() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    index = 1
    for scope in ("region", "school"):
        for combo in logical_combinations():
            rows.append({
                "combination_id": f"combo_{index:03d}",
                "page_scope": scope,
                "학년표현": str(combo["grade"]),
                "과목표현": str(combo["subject"]),
                "내신사용": str(bool(combo["exam"])).lower(),
                "키워드패턴": keyword_pattern(scope, combo),
                "제목사용": "true", "본문사용": "true", "description사용": "true",
                "활성화": "true", "가중치": "1",
                "비고": str(combo["type"]),
            })
            index += 1
    return rows


def write_csv(path: Path, fields: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, fields: list[str], rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "전체"
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for column, field in enumerate(fields, start=1):
        width = max(
            len(field),
            max((len(str(sheet.cell(row=row, column=column).value or "")) for row in range(2, sheet.max_row + 1)), default=0),
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(width + 3, 60)
    workbook.save(path)


def valid_region_name(value: str) -> bool:
    return bool(value) and not re.search(r"(수학|영어|국영수|초등|중등|고등)$", value)


def balanced_entities(rows: list[dict[str, str]], count: int, name_field: str) -> list[dict[str, str]]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        groups[row["시도"]].append(row)
    for values in groups.values():
        values.sort(key=lambda row: stable_int(row["source_html"]))
    used_names: set[str] = set()
    result: list[dict[str, str]] = []
    keys = sorted(groups)
    positions = Counter()
    while len(result) < count:
        progressed = False
        for key in keys:
            values = groups[key]
            while positions[key] < len(values):
                row = values[positions[key]]
                positions[key] += 1
                name = row[name_field]
                if name not in used_names:
                    used_names.add(name)
                    result.append(row)
                    progressed = True
                    break
            if len(result) == count:
                break
        if not progressed:
            raise ValueError(f"고유한 {name_field} {count}건을 선택할 수 없습니다.")
    return result


def school_combo_for(grade: str, index: int) -> dict[str, object]:
    allowed = ["중등", "중학생"] if grade == "중학교" else ["고등", "고등학생"]
    combos = [
        combo for combo in logical_combinations()
        if not combo["grade"] or combo["grade"] in allowed
    ]
    return combos[index % len(combos)]


def display_keyword(entity: str, combo: dict[str, object]) -> str:
    grade = str(combo["grade"])
    subject = str(combo["subject"])
    exam = bool(combo["exam"])
    if not grade and not subject and not exam:
        return f"{entity}과외"
    parts = [entity]
    if grade:
        parts.append(grade)
    if subject:
        parts.append(subject + (" 내신과외" if exam else "과외"))
    elif exam:
        parts.append("내신과외")
    else:
        parts[-1] = parts[-1] + "과외"
    return " ".join(parts)


def core_terms(entity: str, combo: dict[str, object]) -> list[str]:
    terms = [entity]
    if combo["grade"]:
        terms.append(str(combo["grade"]))
    if combo["subject"]:
        terms.append(str(combo["subject"]))
    if combo["exam"]:
        terms.append("내신")
    terms.append("과외")
    return terms


def pick(pool: list[dict[str, str]], identity: str, kind: str, used: set[str]) -> dict[str, str]:
    start = stable_int(f"{identity}:{kind}") % len(pool)
    for offset in range(len(pool)):
        item = pool[(start + offset) % len(pool)]
        if item["id"] not in used:
            used.add(item["id"])
            return item
    raise ValueError(f"문장 풀 선택 실패: {kind}")


def make_title(
    identity: str,
    entity: str,
    keyword: str,
    combo: dict[str, object],
    patterns: list[dict[str, str]],
    modifiers: list[dict[str, str]],
    endings: list[dict[str, str]],
) -> tuple[str, str, str, str]:
    subject_helpers = {
        "과학": [
            "개념 이해와 탐구 자료 해석 방향을 정리합니다",
            "탐구 과정과 서술형 학습 기준을 살펴봅니다",
            "과학 원리와 문제 적용 순서를 점검합니다",
        ],
        "영수": [
            "영어와 수학의 학습 시간 배분을 점검합니다",
            "과목별 복습과 시험 일정 관리 기준을 정리합니다",
            "영수 학습의 균형과 오답 관리 흐름을 살펴봅니다",
        ],
        "영어": [
            "영어 독해와 문장 해석의 학습 방향을 정리합니다",
            "어휘 복습과 독해 문제 접근 기준을 살펴봅니다",
        ],
        "수학": [
            "수학 개념 이해와 문제풀이 순서를 점검합니다",
            "계산 과정과 오답 관리의 학습 기준을 정리합니다",
        ],
        "국어": [
            "국어 독해와 서술형 답안의 학습 방향을 정리합니다",
            "지문 이해와 답의 근거를 찾는 과정을 살펴봅니다",
        ],
    }
    if combo["exam"]:
        helpers = [
            "시험 범위와 오답 관리 기준을 구체적으로 살펴봅니다",
            "교과서와 수행평가를 연결한 내신 준비 흐름을 점검합니다",
            "서술형과 단원별 복습 순서를 차분하게 정리합니다",
        ]
    else:
        helpers = subject_helpers.get(str(combo["subject"]), [
            "학습 방향과 공부 습관을 함께 점검합니다",
            "개념 이해와 오답 관리의 기준을 정리합니다",
            "과목별 학습 계획과 복습 흐름을 살펴봅니다",
        ])
    candidates: list[tuple[int, str, str, str, str]] = []
    for retry in range(100):
        allowed_patterns = [item for item in patterns if combo["exam"] or "내신" not in item["text"]]
        pattern = allowed_patterns[stable_int(f"{identity}:title_pattern:{retry}") % len(allowed_patterns)]
        allowed_modifiers = [item for item in modifiers if combo["exam"] or "내신" not in item["text"]]
        allowed_endings = [item for item in endings if combo["exam"] or "내신" not in item["text"]]
        modifier = allowed_modifiers[stable_int(f"{identity}:title_modifier:{retry}") % len(allowed_modifiers)]
        ending = allowed_endings[stable_int(f"{identity}:title_ending:{retry}") % len(allowed_endings)]
        helper = helpers[stable_int(f"{identity}:title_helper:{retry}") % len(helpers)]
        variants = [
            f"{keyword} {helper}",
            f"{keyword} {ending['text']}",
            f"{keyword} {modifier['text']} {ending['text']}",
        ]
        for title in variants:
            title = re.sub(r"\s+", " ", title).strip()
            if len(title) < 40:
                additions = ["학습 방향", "시험 준비", "개념 이해", "오답 관리", "공부 습관", "과목별 학습 기준"]
                expanded = title
                for addition in additions:
                    if addition not in expanded:
                        expanded = f"{expanded} {addition}을 함께 살펴봅니다"
                    if len(expanded) >= 40:
                        break
                title = expanded
            if 40 <= len(title) <= 68 and title.startswith(keyword + " ") and title.count(keyword) == 1:
                candidates.append((abs(len(title) - 55), title, pattern["id"], modifier["text"], ending["text"]))
    if not candidates:
        raise ValueError(f"제목 생성 실패: {identity}")
    candidates.sort(key=lambda value: value[0])
    _, title, pattern_id, modifier, ending = candidates[0]
    return title, pattern_id, modifier, ending


def make_content(entity: str, combo: dict[str, object], identity: str, pools: dict[str, list[dict[str, str]]]) -> dict[str, object]:
    keyword = display_keyword(entity, combo)
    title, pattern_id, modifier, ending = make_title(
        identity, entity, keyword, combo,
        pools["title_patterns"], pools["title_modifiers"], pools["title_endings"],
    )
    used: set[str] = set()
    description_items = [
        pick(pools["description_general"], identity, f"description_{index}", used)
        for index in range(4)
    ]
    description = f"{keyword}를 살펴볼 때에는 {description_items[0]['text']} " + " ".join(
        item["text"] for item in description_items[1:]
    )
    if len(description) < 250:
        description += " 학습한 내용과 실제 수행 결과를 비교하면 다음 공부의 우선순위를 더 구체적으로 정할 수 있습니다."
    description = re.sub(r"\s+", " ", description).strip()

    body_items = [pick(pools["openings"], identity, "opening", used)]
    body_items.append(pick(pools["body_general"], identity, "general_1", used))
    grade_pool = GRADE_POOL.get(str(combo["grade"]))
    if grade_pool:
        body_items.append(pick(pools[grade_pool], identity, "grade", used))
    else:
        body_items.append(pick(pools["body_general"], identity, "general_2", used))
    subject_pool = SUBJECT_POOL.get(str(combo["subject"]))
    if subject_pool:
        body_items.append(pick(pools[subject_pool], identity, "subject", used))
    else:
        body_items.append(pick(pools["body_general"], identity, "general_3", used))
    if combo["exam"]:
        body_items.append(pick(pools["body_internal_exam"], identity, "internal_exam", used))
    else:
        body_items.append(pick(pools["body_general"], identity, "general_4", used))
    body_items.append(pick(pools["body_general"], identity, "general_5", used))
    body_items.append(pick(pools["endings"], identity, "ending", used))
    body = f"{keyword}를 알아볼 때에는 {body_items[0]['text']} " + " ".join(
        item["text"] for item in body_items[1:]
    )
    terms = core_terms(entity, combo)
    return {
        "keyword": keyword,
        "compact": re.sub(r"\s+", "", keyword),
        "title": title,
        "description": description,
        "body": body,
        "terms": terms,
        "description_ids": [item["id"] for item in description_items],
        "body_ids": [item["id"] for item in body_items],
        "used_pools": sorted({
            "description_general", "openings", "body_general", "endings",
            *([grade_pool] if grade_pool else []),
            *([subject_pool] if subject_pool else []),
            *(["body_internal_exam"] if combo["exam"] else []),
        }),
        "title_pattern_id": pattern_id,
        "title_modifier": modifier,
        "title_ending": ending,
    }


def prepare(replace: bool = False) -> None:
    targets = [PROFILE_CSV, PROFILE_XLSX, SAMPLE_CSV, SAMPLE_XLSX, BUILD_INPUT]
    existing = [str(path) for path in targets if path.exists()]
    if existing and not replace:
        raise FileExistsError("기존 파일을 덮어쓰지 않습니다:\n" + "\n".join(existing))
    profiles = create_profiles()
    write_csv(PROFILE_CSV, PROFILE_FIELDS, profiles)
    write_xlsx(PROFILE_XLSX, PROFILE_FIELDS, profiles)

    with (DATA / "gongbuup_region_master.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        regions = [
            row for row in csv.DictReader(handle)
            if row["is_valid"] == "true" and valid_region_name(row["지역표시명"])
        ]
    with (DATA / "gongbuup_school_master.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        schools = [row for row in csv.DictReader(handle) if row["is_valid"] == "true"]
    region_entities = balanced_entities(regions, 120, "지역표시명")
    middle = balanced_entities([row for row in schools if row["학교급"] == "중학교"], 40, "학교명")
    high = balanced_entities([row for row in schools if row["학교급"] == "고등학교"], 40, "학교명")

    pool_names = [
        "description_general", "openings", "endings", "body_general", "body_elementary",
        "body_middle", "body_high", "body_korean", "body_english", "body_math",
        "body_science", "body_english_math", "body_internal_exam",
        "title_patterns", "title_modifiers", "title_endings",
    ]
    pools = {name: load_json(name) for name in pool_names}
    region_combos = logical_combinations()
    sample_rows: list[dict[str, object]] = []
    build_rows: list[dict[str, object]] = []
    used_titles: set[str] = set()
    used_slugs: set[str] = set()

    selections: list[tuple[str, dict[str, str], dict[str, object]]] = []
    for index, entity_row in enumerate(region_entities):
        selections.append(("region", entity_row, region_combos[index % len(region_combos)]))
    for index, entity_row in enumerate(middle + high):
        selections.append(("school", entity_row, school_combo_for(entity_row["학교급"], index)))

    for page_type, entity_row, combo in selections:
        entity = entity_row["지역표시명"] if page_type == "region" else entity_row["학교명"]
        identity = "|".join((
            page_type, entity_row["시도"], entity_row["시군구"], entity,
            str(combo["grade"]), str(combo["subject"]), str(combo["exam"]),
        ))
        generated = make_content(entity, combo, identity, pools)
        title = str(generated["title"])
        retry = 0
        while title in used_titles and retry < 100:
            retry += 1
            generated = make_content(entity, combo, f"{identity}:title:retry_{retry}", pools)
            title = str(generated["title"])
        slug = str(generated["compact"])
        validation: list[str] = []
        if slug in used_slugs:
            validation.append("슬러그 충돌")
        used_slugs.add(slug)
        used_titles.add(title)
        terms = list(generated["terms"])
        missing_body = [term for term in terms if term not in str(generated["body"])]
        missing_description = [term for term in terms if term not in str(generated["description"])]
        if missing_body:
            validation.append("본문 핵심어 누락: " + ", ".join(missing_body))
        if missing_description:
            validation.append("description 핵심어 누락: " + ", ".join(missing_description))
        if combo["exam"] and str(combo["grade"]) in ("초등", "초등학생"):
            validation.append("초등 내신 금지 조합")
        if not (250 <= len(str(generated["description"])) <= 500):
            validation.append("description 길이 오류")

        terms_text = "|".join(terms)
        sample_rows.append({
            "page_type": page_type,
            "지역명": entity if page_type == "region" else "",
            "학교명": entity if page_type == "school" else "",
            "학교급": entity_row.get("학교급", ""),
            "학년표현": combo["grade"], "과목표현": combo["subject"],
            "내신사용": str(bool(combo["exam"])).lower(),
            "display_keyword": generated["keyword"], "compact_keyword": generated["compact"],
            "slug": slug, "title": title, "title_core_terms": terms_text,
            "body_core_terms": terms_text, "description_core_terms": terms_text,
            "is_valid": "false" if validation else "true",
            "validation_note": "; ".join(validation) or "정상",
        })
        target = (
            "중학생" if entity_row.get("학교급") == "중학교"
            else "고등학생" if entity_row.get("학교급") == "고등학교"
            else TARGET_BY_EXPRESSION.get(str(combo["grade"]), ["초등학생", "중학생", "고등학생"][stable_int(identity) % 3])
        )
        build_rows.append({
            "content_mode": "keyword_combination", "page_type": page_type,
            "시도": entity_row["시도"], "시군구": entity_row["시군구"],
            "읍면동": entity_row.get("읍면동", ""), "리": entity_row.get("리", ""),
            "지역명": entity, "학교명": entity if page_type == "school" else "",
            "학교급": entity_row.get("학교급", ""), "학년표현": combo["grade"],
            "과목표현": combo["subject"], "내신사용": str(bool(combo["exam"])).lower(),
            "대상": target, "과목": str(combo["subject"]) or "국영수과외",
            "메인키워드": generated["keyword"], "슬러그": slug,
            "제목": title, "설명": generated["description"], "본문": generated["body"],
            "body_sentence_ids": "|".join(generated["body_ids"]),
            "description_sentence_ids": "|".join(generated["description_ids"]),
            "used_sentence_pools": "|".join(generated["used_pools"]),
            "title_pattern_id": generated["title_pattern_id"],
            "title_modifier": generated["title_modifier"], "title_ending": generated["title_ending"],
            "combination_type": combo["type"], "title_core_terms": terms_text,
            "title_retry_count": str(retry),
        })
    write_csv(SAMPLE_CSV, SAMPLE_FIELDS, sample_rows)
    write_xlsx(SAMPLE_XLSX, SAMPLE_FIELDS, sample_rows)
    write_csv(BUILD_INPUT, list(build_rows[0]), build_rows)
    print(json.dumps({
        "profiles": len(profiles), "region": 120, "school": 80, "total": 200,
        "invalid": sum(row["is_valid"] != "true" for row in sample_rows),
        "build_input": str(BUILD_INPUT),
    }, ensure_ascii=False, indent=2))


def validate(output_dir: Path) -> None:
    with SAMPLE_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        sample_rows = list(csv.DictReader(handle))
    with BUILD_INPUT.open("r", encoding="utf-8-sig", newline="") as handle:
        build_rows = list(csv.DictReader(handle))
    build_by_slug = {row["슬러그"]: row for row in build_rows}
    counts = Counter()
    titles: list[str] = []
    descriptions: list[str] = []
    bodies: list[str] = []
    slugs: list[str] = []

    for sample in sample_rows:
        slug = sample["slug"]
        build = build_by_slug[slug]
        page = output_dir / "과외" / slug / "index.html"
        notes: list[str] = []
        if not page.is_file():
            notes.append("HTML 누락")
            counts["html_missing"] += 1
            continue
        text = page.read_text(encoding="utf-8-sig")
        title = html.unescape(re.search(r"<title>(.*?)</title>", text, re.S).group(1))
        description = html.unescape(re.search(r'<meta name="description" content="([^"]*)"', text).group(1))
        body_html = re.search(r"<p>(.*?)</p>", text, re.S).group(1)
        body = html.unescape(re.sub(r"<[^>]+>", " ", body_html))
        h1 = html.unescape(re.search(r"<h1>(.*?)</h1>", text, re.S).group(1))
        terms = sample["title_core_terms"].split("|")
        missing_title = [term for term in terms if term not in title]
        missing_body = [term for term in terms if term not in body]
        missing_description = [term for term in terms if term not in description]
        counts["required_title_term_missing"] += int(bool(missing_title))
        counts["body_core_missing"] += int(bool(missing_body))
        counts["description_core_missing"] += int(bool(missing_description))
        if missing_title:
            notes.append("제목 필수 키워드 누락")
        if missing_body:
            notes.append("본문 핵심어 누락")
        if missing_description:
            notes.append("description 핵심어 누락")
        if h1 != sample["display_keyword"]:
            notes.append("h1 불일치")
            counts["h1_error"] += 1
        if sample["내신사용"] == "true" and "내신" not in title:
            counts["exam_mismatch"] += 1
        if sample["내신사용"] == "false" and "내신" in title:
            counts["exam_mismatch"] += 1
        if sample["학년표현"] and sample["학년표현"] not in title:
            counts["target_mismatch"] += 1
        if sample["과목표현"] and sample["과목표현"] not in title:
            counts["subject_mismatch"] += 1
        canonical = re.search(r'<link rel="canonical" href="([^"]+)"', text)
        og_image = re.search(r'<meta property="og:image" content="([^"]+)"', text)
        if not canonical or not canonical.group(1).endswith(f"/{quote_path('과외')}/{quote_path(slug)}/"):
            counts["canonical_error"] += 1
        if not og_image:
            counts["image_error"] += 1
        else:
            relative = unquote(urlparse(og_image.group(1)).path).lstrip("/")
            if not (output_dir / Path(relative.replace("/", "\\"))).is_file():
                counts["image_error"] += 1
        for source in re.findall(r'<img[^>]+src="([^"]+)"', text):
            relative = unquote(urlparse(source).path).lstrip("/")
            if not (output_dir / Path(relative.replace("/", "\\"))).is_file():
                counts["image_error"] += 1
        sample["is_valid"] = "false" if notes else "true"
        sample["validation_note"] = "; ".join(notes) or "정상"
        titles.append(title)
        descriptions.append(description)
        bodies.append(body)
        slugs.append(slug)

    counts["title_duplicate"] = len(titles) - len(set(titles))
    counts["description_duplicate"] = len(descriptions) - len(set(descriptions))
    counts["body_duplicate"] = len(bodies) - len(set(bodies))
    counts["slug_collision"] = len(slugs) - len(set(slugs))
    title_lengths = [len(title) for title in titles]
    title_retry_counts = [int(row.get("title_retry_count", "0") or 0) for row in build_rows]
    combo_counts = Counter(row["combination_type"] for row in build_rows)
    subject_counts = Counter(row["과목표현"] for row in build_rows if row["과목표현"])
    science_pool_missing = sum(
        row["과목표현"] == "과학" and "body_science" not in row["used_sentence_pools"].split("|")
        for row in build_rows
    )
    korean_pool_missing = sum(
        row["과목표현"] == "국어" and "body_korean" not in row["used_sentence_pools"].split("|")
        for row in build_rows
    )
    english_math_pool_missing = sum(
        row["과목표현"] == "영수" and "body_english_math" not in row["used_sentence_pools"].split("|")
        for row in build_rows
    )
    elementary_exam = sum(
        row["내신사용"] == "true" and row["학년표현"] == "초등" for row in build_rows
    )
    elementary_student_exam = sum(
        row["내신사용"] == "true" and row["학년표현"] == "초등학생" for row in build_rows
    )
    write_csv(SAMPLE_CSV, SAMPLE_FIELDS, sample_rows)
    write_xlsx(SAMPLE_XLSX, SAMPLE_FIELDS, sample_rows)
    generated_files = [
        PROFILE_CSV, PROFILE_XLSX, SAMPLE_CSV, SAMPLE_XLSX, BUILD_INPUT, REPORT_JSON, REPORT_TXT,
        POOLS / "body_science.json", POOLS / "body_english_math.json", POOLS / "body_internal_exam.json",
    ]
    report = {
        "output_dir": str(output_dir.relative_to(ROOT)),
        "total_sample_page_count": len(build_rows),
        "region_page_count": sum(row["page_type"] == "region" for row in build_rows),
        "school_page_count": sum(row["page_type"] == "school" for row in build_rows),
        "combination_type_count": dict(combo_counts),
        "middle_or_middle_student_exam_count": sum(
            row["내신사용"] == "true" and row["학년표현"] in ("중등", "중학생") for row in build_rows
        ),
        "high_or_high_student_exam_count": sum(
            row["내신사용"] == "true" and row["학년표현"] in ("고등", "고등학생") for row in build_rows
        ),
        "elementary_exam_count": elementary_exam,
        "elementary_student_exam_count": elementary_student_exam,
        "subject_usage_count": dict(subject_counts),
        "body_core_term_missing_count": counts["body_core_missing"],
        "description_core_term_missing_count": counts["description_core_missing"],
        "title_duplicate_count": counts["title_duplicate"],
        "title_minimum_length": min(title_lengths, default=0),
        "title_maximum_length": max(title_lengths, default=0),
        "title_average_length": round(sum(title_lengths) / len(title_lengths), 2) if title_lengths else 0,
        "title_below_minimum_count": sum(length < 40 for length in title_lengths),
        "title_above_maximum_count": sum(length > 68 for length in title_lengths),
        "title_regenerated_page_count": sum(count > 0 for count in title_retry_counts),
        "title_max_retry_count": max(title_retry_counts, default=0),
        "required_title_term_missing_count": counts["required_title_term_missing"],
        "description_duplicate_count": counts["description_duplicate"],
        "body_duplicate_count": counts["body_duplicate"],
        "slug_collision_count": counts["slug_collision"],
        "target_mismatch_count": counts["target_mismatch"],
        "subject_mismatch_count": counts["subject_mismatch"],
        "internal_exam_mismatch_count": counts["exam_mismatch"],
        "science_pool_not_applied_count": science_pool_missing,
        "korean_pool_not_applied_count": korean_pool_missing,
        "english_math_pool_not_applied_count": english_math_pool_missing,
        "canonical_error_count": counts["canonical_error"],
        "image_error_count": counts["image_error"],
        "generated_files": [str(path.relative_to(ROOT)) for path in generated_files],
        "modified_files": [
            "config/settings.json", "scripts/build.py", "scripts/create_sentence_pools.py",
            "scripts/keyword_combination_engine.py",
        ],
    }
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = [
        "키워드 조합 검증 보고서",
        f"전체 샘플 페이지 수: {report['total_sample_page_count']}",
        f"지역 페이지 수: {report['region_page_count']}",
        f"학교 페이지 수: {report['school_page_count']}",
        *[f"{name} 조합 수: {combo_counts.get(name, 0)}" for name in (
            "base", "subject_only", "grade_only", "grade_subject", "exam_only",
            "subject_exam", "grade_exam", "grade_subject_exam",
        )],
        f"중등/중학생 + 내신 조합 수: {report['middle_or_middle_student_exam_count']}",
        f"고등/고등학생 + 내신 조합 수: {report['high_or_high_student_exam_count']}",
        f"초등 + 내신 조합 수: {elementary_exam}",
        f"초등학생 + 내신 조합 수: {elementary_student_exam}",
        *[f"{subject} 사용 수: {subject_counts.get(subject, 0)}" for subject in SUBJECTS],
        f"제목 핵심 키워드 본문 누락 수: {counts['body_core_missing']}",
        f"제목 핵심 키워드 description 누락 수: {counts['description_core_missing']}",
        f"title 중복 수: {counts['title_duplicate']}",
        f"제목 최소 길이: {report['title_minimum_length']}",
        f"제목 최대 길이: {report['title_maximum_length']}",
        f"제목 평균 길이: {report['title_average_length']}",
        f"40자 미만 제목 수: {report['title_below_minimum_count']}",
        f"68자 초과 제목 수: {report['title_above_maximum_count']}",
        f"title 재생성 실행 페이지 수: {report['title_regenerated_page_count']}",
        f"최대 재시도 횟수: {report['title_max_retry_count']}",
        f"필수 키워드 누락 수: {counts['required_title_term_missing']}",
        f"description 중복 수: {counts['description_duplicate']}",
        f"본문 중복 수: {counts['body_duplicate']}",
        f"슬러그 충돌 수: {counts['slug_collision']}",
        f"대상 불일치 수: {counts['target_mismatch']}",
        f"과목 불일치 수: {counts['subject_mismatch']}",
        f"내신 불일치 수: {counts['exam_mismatch']}",
        f"과학 문장 풀 미적용 수: {science_pool_missing}",
        f"국어 문장 풀 미적용 수: {korean_pool_missing}",
        f"영수 문장 풀 미적용 수: {english_math_pool_missing}",
        f"canonical 오류 수: {counts['canonical_error']}",
        f"이미지 오류 수: {counts['image_error']}",
        "",
        "생성한 파일:",
        *[f"- {path}" for path in report["generated_files"]],
        "",
        "수정한 파일:",
        *[f"- {path}" for path in report["modified_files"]],
    ]
    REPORT_TXT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


def quote_path(value: str) -> str:
    from urllib.parse import quote
    return quote(value, safe="")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=("prepare", "validate"))
    parser.add_argument("--output-dir")
    parser.add_argument("--replace", action="store_true")
    args = parser.parse_args()
    if args.action == "prepare":
        prepare(replace=args.replace)
    else:
        if not args.output_dir:
            raise ValueError("validate에는 --output-dir이 필요합니다.")
        output_dir = Path(args.output_dir)
        if not output_dir.is_absolute():
            output_dir = ROOT / output_dir
        validate(output_dir)


if __name__ == "__main__":
    main()
