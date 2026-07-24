from __future__ import annotations

import csv
import json
import os
import tempfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
REPORTS = ROOT / "reports"
OLD_SOURCE_NAME = "이화여자대학교사범대학부속이화금란중학교"
OFFICIAL_NAME = "이화여자대학교사범대학부속이화·금란중학교"


def update_csv(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        fields = list(reader.fieldnames or [])
        rows = list(reader)
    changed = 0
    for row in rows:
        if row.get("학교명") == OLD_SOURCE_NAME:
            row["학교명"] = OFFICIAL_NAME
            changed += 1
    if changed != 1:
        raise RuntimeError(f"{path}: expected one school row, found {changed}")
    handle, temp_name = tempfile.mkstemp(
        prefix=path.stem + "_", suffix=".csv", dir=path.parent
    )
    os.close(handle)
    temp = Path(temp_name)
    try:
        with temp.open("w", encoding="utf-8-sig", newline="") as target:
            writer = csv.DictWriter(target, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
        os.replace(temp, path)
    finally:
        temp.unlink(missing_ok=True)
    return changed


def update_xlsx(path: Path) -> int:
    workbook = load_workbook(path)
    changed = 0
    for sheet in workbook.worksheets:
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None
        }
        column = headers.get("학교명")
        if not column:
            continue
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            if cell.value == OLD_SOURCE_NAME:
                cell.value = OFFICIAL_NAME
                changed += 1
    if changed != 1:
        workbook.close()
        raise RuntimeError(f"{path}: expected one school row, found {changed}")
    workbook.save(path)
    workbook.close()
    return changed


def main() -> int:
    csv_path = DATA / "gongbuup_school_master.csv"
    xlsx_path = DATA / "학교명_전체.xlsx"
    result = {
        "official_school_name": OFFICIAL_NAME,
        "csv_changed_rows": update_csv(csv_path),
        "xlsx_changed_rows": update_xlsx(xlsx_path),
        "files": [str(csv_path), str(xlsx_path)],
    }
    REPORTS.mkdir(parents=True, exist_ok=True)
    report = REPORTS / "ewha_geumran_school_correction.json"
    report.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
